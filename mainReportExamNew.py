#cào báo cáo dữ liệu khám
#using
import os
import time
import tkinter as tk
from tkinter import messagebox
from datetime import datetime
import customtkinter
from tkinter import filedialog
from seleniumwire import webdriver
from selenium.webdriver.common.by import By
from PIL import ImageTk, Image
from tkcalendar import Calendar
from babel.numbers import *
import json
import pandas as pd
from openpyxl.cell import MergedCell
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
import threading
import sourceString as sour
import requests
from collections import Counter

#global
customtkinter.set_appearance_mode("Light")
customtkinter.set_default_color_theme("green")
listStatus = []
current_date = datetime.datetime.today()
date_entry = Any
date_entry2 = Any
run_button = Any
app = Any
urlAPI = "http://192.168.0.77/api/out_patient_new_regist/baoCaoSoKhamBenh"
list_big = []

#class
class PatientModel:
    stt = 0
    code = ""
    dateExam = ""
    status = ""
    prescription = ""
    doctor = ""
    ticket = ""

    def to_dict(self):
        return {
            "STT": self.stt,
            "Mã bệnh nhân": self.code,
            "Ngày khám": self.dateExam,
            "Mã toa thuốc": self.prescription,
            "Tình trạng": self.status,
            "Bác sĩ": self.doctor,
            "Mã đợt khám": self.ticket
        }

    def printModel(self):
        print("-------------------------------------------------")
        print(f"STT: {self.stt}")
        print(f"Mã bệnh nhân: {self.code}")
        print(f"Ngày khám: {self.dateExam}")
        print(f"Mã toa thuốc: {self.prescription}")
        print(f"Tình trạng: {self.status}")
        print(f"Bác sĩ: {self.doctor}")
        print("-------------------------------------------------")

    def ExportModel(self):
        headers = [
            ("STT:", self.stt),
            ("Mã bệnh nhân:", self.code),
            ("Ngày khám:", self.dateExam),
            ("Mã toa thuốc:", self.prescription),
            ("Tình trạng:", self.status),
            ("Bác sĩ:", self.doctor),
            ("Mã đợt khám:", self.ticket)
        ]

        # Đảm bảo tất cả các giá trị không phải là None
        headers = [(header, "" if value is None else value) for header, value in headers]

        # Tính toán độ rộng cột
        max_key_length = max(len(header[0]) for header in headers) + 2
        max_value_length = max(len(str(header[1])) for header in headers) + 2

        s = ""
        s += "+" + "-" * max_key_length + "+" + "-" * max_value_length + "+\n"

        for header, value in headers:
            key_column = f" {header:<{max_key_length-2}} "
            value_column = f" {value:<{max_value_length-2}} "
            s += f"|{key_column}|{value_column}|\n"

        s += "+" + "-" * max_key_length + "+" + "-" * max_value_length + "+\n"
        return s

class StatusExamModel:
    # thuộc tính
    status_code = 0
    status_name = ""
    # phương thức
    def setName(self, status_name):
        self.status_name = status_name
    
    def getName(self):
        return self.status_name
    
    def setCode(self, status_code):
        self.status_code = status_code
    
    def getCode(self):
        return self.status_code
    
class DatePicker(tk.Toplevel):
    def __init__(self, parent, date_entry, next_widget):
        super().__init__(parent)
        self.title("Date Picker")
        self.overrideredirect(True)
        self.grab_set()
        self.date_entry = date_entry
        self.next_widget = next_widget

        self.cal = Calendar(self, selectmode="day", year=current_date.year, month=current_date.month,
                            day=current_date.day)
        self.cal.pack()

        confirm_button = customtkinter.CTkButton(self, text="Confirm", command=self.update_date)
        confirm_button.pack()

    def update_date(self):
        selected_date_str = self.cal.get_date()
        selected_date = datetime.datetime.strptime(selected_date_str, "%m/%d/%y")  # Convert to datetime object
        self.date_entry.delete(0, tk.END)
        self.date_entry.insert(0, selected_date.strftime("%d/%m/%Y"))
        self.destroy()
        self.next_widget.focus_set()

def loadingData():
    s1 = StatusExamModel()
    s2 = StatusExamModel()
    s3 = StatusExamModel()
    s4 = StatusExamModel()
    s1.setCode(0)
    s1.setName("Chưa khám")
    s2.setCode(1)
    s2.setName("Đã khám")
    s3.setCode(3)
    s3.setName("Chờ CLS")
    s4.setCode(5)
    s4.setName("")
    listStatus.append(s1)
    listStatus.append(s2)
    listStatus.append(s3)
    listStatus.append(s4)

def convert_date_format(date_str):
    try:
        date_obj = datetime.datetime.strptime(date_str, "%d/%m/%Y")
        return date_obj.strftime("%Y-%m-%d")
    except ValueError:
        print(f"Invalid date format: {date_str}")
        return date_str

def pick_date(event):
    next_widget = date_entry2
    date_picker = DatePicker(app, date_entry, next_widget)
    x, y, _, _ = event.widget.bbox("insert")
    x += event.widget.winfo_rootx() - 7
    y += event.widget.winfo_rooty() + 30
    date_picker.geometry("+%d+%d" % (x, y))

def pick_date2(event):
    next_widget = run_button
    date_picker = DatePicker(app, date_entry2, next_widget)
    x, y, _, _ = event.widget.bbox("insert")
    x += event.widget.winfo_rootx() - 7
    y += event.widget.winfo_rooty() + 30
    date_picker.geometry("+%d+%d" % (x, y))

def validate_dates():
    date_format = "%d/%m/%Y"
    date1_str = date_entry.get()
    date2_str = date_entry2.get()

    try:
        date1 = datetime.datetime.strptime(date1_str, date_format)
        date2 = datetime.datetime.strptime(date2_str, date_format)

        if date2 < date1:
            return False
        
        return True
    except ValueError:
        return False

def is_in_icd_range(icd_code, icd_range):
    # Kiểm tra khớp chính xác
    if icd_code in icd_range:
        return True
    
    if icd_code == None:
            return False

    # Kiểm tra khớp với mã gốc chỉ khi mã đầu vào không có phần thập phân
    if '.' not in icd_code:
        return any(code.startswith(icd_code + '.') for code in icd_range)
    
    return False

# Danh sách mã ICD-10
icd_range = [
    "N18",
    "N18.0",
    "N18.1",
    "N18.2",
    "N18.3",
    "N18.4",
    "N18.5",
    "D69",
    "D69.0",
    "M32",
    "N04",
    "N32"
]

def count_icd_occurrences(models, icd_range):
    # Tạo một Counter để đếm số lần xuất hiện
    icd_counter = Counter()

    for model in models:
        primary_icd = model['primary_icd10_code']
        
        # Kiểm tra xem primary_icd có khớp với bất kỳ mã nào trong icd_range không
        for icd in icd_range:
            if primary_icd == icd or (not '.' in icd and primary_icd.startswith(icd + '.')):
                icd_counter[icd] += 1
                break  # Dừng sau khi tìm thấy khớp đầu tiên

    # Chuyển đổi Counter thành dict, bao gồm cả các mã có số đếm là 0
    result = {icd: icd_counter[icd] for icd in icd_range}
    
    return result

def save_to_json(data, filename):
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

def open_terminal_window():
    terminal_window = tk.Toplevel(app)
    terminal_window.title("Màn hình chạy dữ liệu")
    terminal_window.attributes("-fullscreen", True)
    icon_path = os.path.abspath("resource\crawlLogo.ico")
    terminal_window.iconbitmap(icon_path)
    terminal_text = tk.Text(terminal_window, bg="black", fg="green", insertbackground="green")
    terminal_text.pack(expand=True, fill='both')

    def on_closing():
        if script_thread.is_alive():
            terminal_window.destroy()
            app.deiconify()
        else:
            terminal_window.destroy()
            app.deiconify()

    terminal_window.protocol("WM_DELETE_WINDOW", on_closing)
    return terminal_window, terminal_text

def start_script_thread():
    global script_thread  # Khai báo biến toàn cục
    directory = filedialog.askdirectory(title="Chọn nơi lưu trữ!")
    if directory:
        directory = directory.replace('/', '\\')
    else:
        return
    terminal_window, terminal_text = open_terminal_window()
    app.withdraw()
    script_thread = threading.Thread(target=run_Script, args=(directory, terminal_text))
    script_thread.start()
# Hàm định dạng sheet
def center_window(window, width=600, height=440):
    # Lấy kích thước màn hình
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    # Tính toán tọa độ x và y để cửa sổ ở giữa màn hình
    x = (screen_width / 2) - (width / 2)
    y = (screen_height / 2) - (height / 2)
    window.geometry(f'{width}x{height}+{int(x)}+{int(y)}')
    
def search_in_list_big(model):
    global list_big
    if len(list_big) > 0:
        for item in list_big:
            paCode = item['patient_code']
            icdCode = item['primary_icd10_code']
            if paCode == model['patient_code']:
                if icdCode == model['primary_icd10_code']:
                    return True
        return False
    else:
        return False
def format_sheet(ws, date_str):
    # Dòng tiêu đề chính
    ws.merge_cells('A1:W1')
    title_cell = ws['A1']
    title_cell.value = "Báo cáo sổ khám bệnh"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal="center")

    # Dòng ngày khám
    ws.merge_cells('A2:W2')
    date_cell = ws['A2']
    date_cell.value = f"Ngày khám: {date_str}"
    date_cell.font = Font(size=14, bold=True)
    date_cell.alignment = Alignment(horizontal="center")

    # Dòng header các cột
    headers = ["STT", "ID bệnh nhân", "Mã bệnh nhân", "out_patient_list_id", "Tên bệnh nhân", "Giới tính", "Ngày sinh", "Địa chỉ", "Code thân nhân", "Trạng thái", "Ngày khám", "Mã đợt khám", "Mã hàng chờ", "Ca", "Loại khám", "Mã khu vực", "BHLS tạm", "Chuẩn đoán tạm", "BHLS", "Chuẩn đoán", "Tên bác sĩ", "Mã ICD", "Tên ICD"]
    header_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = Font(size=13, bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = header_fill

    # Tự động điều chỉnh độ rộng cột
    for col in ws.iter_cols(min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        max_length = 0
        column = col[0].column_letter  # Lấy tên cột
        for cell in col:
            if isinstance(cell, MergedCell):
                continue  # Bỏ qua ô đã hợp nhất
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

def find_status(id):
    for s in listStatus:
        if s.getCode() == id:
            return s.getName()
    return None

def run_Script(directory,terminal_text):
    global  date_entry, date_entry2, app, current_date, list_big
    # Hàm để hiển thị thông tin lên terminal
    def log_terminal(message):
        terminal_text.insert(tk.END,message + "\n")
        terminal_text.see(tk.END)  # Scroll to the end
        terminal_text.update_idletasks()
    log_terminal(".........................................................................................")
    log_terminal(".........................Kiểm tra thời gian thực thi.....................................")

    if validate_dates() == True:
        log_terminal(".........................Thời gian hợp lệ................................................")
        start_date = datetime.datetime.strptime(date_entry.get(), "%d/%m/%Y")
        end_date = datetime.datetime.strptime(date_entry2.get(), "%d/%m/%Y")
        date_list = []
        current_date = start_date

        while current_date <= end_date:
            date_list.append(current_date)
            current_date += datetime.timedelta(days=1)
        
        log_terminal(".........................Xác định ngày thành công........................................")
        file_name = f"Dữ liệu khám {start_date.strftime('%d-%m-%Y')}_{end_date.strftime('%d-%m-%Y')}.xlsx"
        file_path = os.path.join(directory, file_name)
        # Tạo hoặc mở file Excel
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
        else:
            wb = Workbook()
            wb.remove(wb.active)  # Xóa sheet mặc định khi tạo mới
        log_terminal(".........................Tạo file Excel lưu thành công...................................")
        log_terminal(".........................Khởi động Chrome................................................")
        #chrome settings
        bTry = False
        errorChrome = 1
        while bTry == False:
            bTry = sour._initSelenium_()
            if bTry == False:
                if errorChrome >= 10:
                    log_terminal("..............Khởi động Chrome thất bại quá nhiều! Tắt chương trình....................")
                    app.destroy() 
                else:
                    errorChrome += 1
                    log_terminal(".........................Khởi động Chrome thất bại! Đang thử lại.......................")
        
        driver = sour.driverGlobal
        time.sleep(3)
        log_terminal(".........................Duyệt WEBSITE thành công........................................")
        # login
        sour._login_("quyen.ngoq", "74777477")
        # ấn nút login
        time.sleep(0.5)
        for date in date_list:
            headers = {
                        "Appkey": sour.Appkey,
                        "Userkey": sour.secretKey,
                        "Authorization": sour.secretKey,
                        "Content-Type": sour.contentType
                    }
            datefm = date.strftime("%Y-%m-%d")
            params = {
                        "date": datefm,
                        "zoneCode": "nguyendu",
                        "shift": "-1"
                    }
            demstt = 1
            listPatient = []
            response = requests.get(urlAPI, headers=headers, params=params)
            if response.status_code == 200:
                # Nếu request thành công, lấy dữ liệu JSON
                data = response.json()
                patient_info = data.get('data', {})
                for pi in patient_info:
                    result = is_in_icd_range(pi.get('primary_icd10_code'), icd_range)
                    if result == True:
                        # patientmd = {
                        #     "stt": demstt,
                        #     "patient_id": pi.get('patient_id'),
                        #     "patient_code": pi.get('patient_code'),
                        #     "out_patient_list_id": pi.get('out_patient_list_id'),
                        #     "name": pi.get('name'),
                        #     "gender": pi.get('gender'),
                        #     "date_of_birth": pi.get('date_of_birth'),
                        #     "full_address": pi.get('full_address'),
                        #     "relative_name": pi.get('relative_name'),
                        #     "status": pi.get('status'),
                        #     "register_date": date,
                        #     "ticket_id": pi.get('ticket_id'),
                        #     "queue_code": pi.get('queue_code'),
                        #     "shift": pi.get('shift'),
                        #     "type_exam": pi.get('type_exam'),
                        #     "zone_code": pi.get('zone_code'),
                        #     "bhls_temp": pi.get('bhls_temp'),
                        #     "chandoan_temp": pi.get('chandoan_temp'),
                        #     "bhls": pi.get('bhls'),
                        #     "chandoan": pi.get('chandoan'),
                        #     "doctor_name": pi.get('doctor_name'),
                        #     "primary_icd10_code": pi.get('primary_icd10_code'),
                        #     "primary_icd10_name": pi.get('primary_icd10_name')
                        # }
                        patientmd = {
                            "patient_code": pi.get('patient_code'),
                            "primary_icd10_code": pi.get('primary_icd10_code')
                        }
                        listPatient.append(patientmd)
                        if search_in_list_big(patientmd) == False:
                            list_big.append(patientmd)
                        #log_terminal(patientmd.ExportModel())
                        demstt+=1
            else:
                print(f"Request thất bại với status code: {response.status_code}")
                print(response.text)
            # Chuyển listPatient thành DataFrame
            #patient_dict_list = [pt.to_dict() for pt in listPatient]
            # df = pd.DataFrame(listPatient)
            # log_terminal(".........................Chuyển dữ liệu thành công.......................................")
            # # Tạo sheet mới cho từng ngày
            # sheet_name = date.strftime("%d-%m-%Y")
            # ws = wb.create_sheet(title=sheet_name)
            # log_terminal(".........................Tạo Sheet thành công............................................")
            # # Thêm dữ liệu vào sheet
            # for r in range(3):
            #     ws.append([])

            # for r in dataframe_to_rows(df, index=False, header=False):
            #     ws.append(r)

            # # Định dạng sheet
            # format_sheet(ws, sheet_name)
            # log_terminal(".........................Định dạng sheet thành công......................................")
        # Lưu file Excel
        # wb.save(file_path)
        # Đếm số lần xuất hiện
        result2 = count_icd_occurrences(list_big, icd_range)

        # Lưu kết quả vào file JSON
        save_to_json(result2, 'icd_occurrences.json')
        log_terminal(".........................Hoàn thành xuất dữ liệu báo cáo.................................")
        app.deiconify()
    else:
        messagebox.showerror(title="Lỗi thời gian",message="Thời gian kết thúc phải lớn hơn hoặc bằng bắt đầu!")
        app.deiconify()

#app
def run_secondary_interface(main_app):
    global run_button, date_entry, date_entry2, app, current_date
    app = customtkinter.CTkToplevel(main_app)
    app.title("Lấy báo cáo sổ khám bệnh")
    center_window(app)
    icon_path = os.path.abspath("resource\crawlLogo.ico")
    app.iconbitmap(icon_path)

    def on_closing():
        main_app.deiconify()  # Hiển thị lại cửa sổ chính khi đóng cửa sổ mới
        app.destroy()

    app.protocol("WM_DELETE_WINDOW", on_closing)

    imgBG = ImageTk.PhotoImage(Image.open("resource\BG2.jpg"))
    l1 = customtkinter.CTkLabel(master=app, image=imgBG)
    l1.pack()

    frame = customtkinter.CTkFrame(master=l1, width=320, height=250, corner_radius=15)
    frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

    l2 = customtkinter.CTkLabel(master=frame, text="Select configuration", font=('Century Gothic', 20))
    l2.place(x=70, y=45)

    loadingData()
    
    l3 = customtkinter.CTkLabel(master=frame, text="Từ ngày", font=('Roboto', 13))
    l3.place(x=20, y=90)
    date_entry = customtkinter.CTkEntry(master=frame, width=200, corner_radius=10)
    date_entry.place(x=100, y=90)
    date_entry.insert(0, current_date.strftime("%d/%m/%Y"))  # Đặt giá trị mặc định là ngày hiện tại
    date_entry.bind("<FocusIn>", pick_date)

    l4 = customtkinter.CTkLabel(master=frame, text="Đến ngày", font=('Roboto', 13))
    l4.place(x=20, y=130)
    date_entry2 = customtkinter.CTkEntry(master=frame, width=200, corner_radius=10)
    date_entry2.place(x=100, y=130)
    date_entry2.insert(0, current_date.strftime("%d/%m/%Y"))  # Đặt giá trị mặc định là ngày hiện tại
    date_entry2.bind("<FocusIn>", pick_date2)

    run_button = customtkinter.CTkButton(master=frame, command=start_script_thread,text="Thực thi", font=('Tahoma', 13), fg_color="#005369", hover_color="#008097")
    run_button.place(x=160, y=200)

    app.mainloop()