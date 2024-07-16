import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import re
from datetime import datetime

def date_converter_with_invalid(x):
    if str(x) == "0000-00-00":
        return "0000-00-00"
    try:
        x = pd.to_datetime(x)
        return x.strftime('%Y-%m-%d')
    except:
        return str(x)
    
def choose_save_path():
    path = filedialog.askdirectory()
    if path:
        save_path_entry.delete(0, tk.END)
        save_path_entry.insert(0, path)

def date_converter(x):
    if pd.isna(x):
        return ''
    if isinstance(x, datetime):
        return x.strftime('%Y-%m-%d')
    # Nếu không phải là timestamp, sử dụng regex để trích xuất ngày tháng năm
    date_match = re.search(r'(\d{4}-\d{2}-\d{2})', str(x))
    if date_match:
        return date_match.group(1)
    return str(x)
    
def select_files():
    files = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx")])
    if files:
        process_files(files)

def process_files(files):
    output_name = output_name_entry.get()
    save_path = save_path_entry.get()
    
    if not output_name or not save_path:
        messagebox.showerror("Error", "Please enter output file name and select save path.")
        return
    
    output_file = os.path.join(save_path, f"{output_name}.xlsx")
    
    # Tạo một workbook mới
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Xóa sheet mặc định

    for file in files:
        # Lấy ngày từ tên file
        date = os.path.basename(file).split('_')[-1].split('.')[0]
        
        # Đọc file Excel gốc
        df = pd.read_excel(file, engine='openpyxl', dtype={1: str})
        
        # Xử lý cột thứ 4 và thứ 10 thành dạng ngày tháng năm
        df.iloc[:, 3] = df.iloc[:, 3].apply(date_converter_with_invalid)
        df.iloc[:, 9] = df.iloc[:, 9].apply(date_converter_with_invalid)
        
        # Tạo sheet mới với tên là ngày
        ws = wb.create_sheet(title=date)
        
        # Sao chép dữ liệu từ DataFrame vào worksheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Định dạng header
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        header_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        header_border = Border(
            left=Side(style='thick'),
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )
        
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill
            cell.border = header_border

        for col_idx in [4, 10]:  # Cột 4 và 10 (index từ 1)
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.number_format = 'YYYY-MM-DD'

        # Sao chép định dạng từ file gốc
        src_wb = openpyxl.load_workbook(file)
        src_ws = src_wb.active
        
        # Sao chép chiều rộng cột và chiều cao hàng
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = src_ws.column_dimensions[col[0].column_letter].width
        for row in ws.rows:
            ws.row_dimensions[row[0].row].height = src_ws.row_dimensions[row[0].row].height

    # Lưu workbook
    wb.save(output_file)
    
    messagebox.showinfo("Success", f"Files combined successfully. Output saved as {output_file}")

# Tạo cửa sổ chính
root = tk.Tk()
root.title("Excel File Combiner")

# Tạo và đặt các widget
tk.Label(root, text="Output File Name:").grid(row=0, column=0, sticky="e", padx=5, pady=5)
output_name_entry = tk.Entry(root, width=40)
output_name_entry.grid(row=0, column=1, columnspan=2, padx=5, pady=5)

tk.Label(root, text="Save Path:").grid(row=1, column=0, sticky="e", padx=5, pady=5)
save_path_entry = tk.Entry(root, width=40)
save_path_entry.grid(row=1, column=1, padx=5, pady=5)
tk.Button(root, text="Browse", command=choose_save_path).grid(row=1, column=2, padx=5, pady=5)

tk.Button(root, text="Select and Combine Files", command=select_files).grid(row=2, column=0, columnspan=3, pady=10)

root.mainloop()
