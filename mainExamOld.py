#Cào dữ liệu khám có database
#using
import threading
from tkinter import Menu
import os
import tkinter as tk
from tkinter import messagebox
from datetime import datetime
import customtkinter
from tkinter import filedialog
from PIL import ImageTk, Image
from tkcalendar import Calendar
from babel.numbers import *
import pandas as pd
from openpyxl import Workbook,load_workbook
from tkinter import ttk
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
import mysql.connector as connector

#global
customtkinter.set_appearance_mode("Light")
customtkinter.set_default_color_theme("green")
listStatus = []
current_date = datetime.datetime.today()
date_entry = Any
date_entry2 = Any
run_button = Any
app = Any
#class
class PatientModel:
    stt = 0
    code = ""
    name = ""
    datebirth = ""
    gender = ""
    age = 0
    province = ""
    district = ""
    ward = ""
    date = ""
    shift = ""
    primaryICD10Code = ""
    primaryICD10 = ""
    secondICD10 = ""
    diagnosis = ""
    status = ""

    def to_dict(self):
        return {
            "STT": self.stt,
            "Mã bệnh nhân": self.code,
            "Tên bệnh nhân": self.name,
            "Ngày sinh": self.datebirth,
            "Giới tính": self.gender,
            "Tuổi": self.age,
            "Tỉnh Thành": self.province,
            "Quận Huyện": self.district,
            "Xã Phường": self.ward,
            "Ngày khám": self.date,
            "Ca khám": self.shift,
            "Mã ICD chính": self.primaryICD10Code,
            "Tên ICD chính": self.primaryICD10,
            "Mã ICD phụ": self.secondICD10,
            "Chuẩn đoán 2": self.diagnosis,
            "Tình trạng": self.status,
        }

    def printModel(self):
        print("-------------------------------------------------")
        print(f"STT: {self.stt}")
        print(f"Mã bệnh nhân: {self.code}")
        print(f"Tên bệnh nhân: {self.name}")
        print(f"Ngày sinh: {self.datebirth}")
        print(f"Tuổi: {self.age}")
        print(f"Giới tính: {self.gender}")
        print(f"Phường/Xã: {self.ward}")
        print(f"Quận/Huyện: {self.district}")
        print(f"Tỉnh/Thành Phố: {self.province}")
        print(f"Ngày khám: {self.date}")
        print(f"Ca Khám: {self.shift}")
        print(f"ICD1: {self.primaryICD10Code}")
        print(f"Tên ICD1: {self.primaryICD10}")
        print(f"ICD 2 (phụ): {self.secondICD10}")
        print(f"Chuẩn đoán 2: {self.diagnosis}")
        print(f"Tình trạng: {self.status}")
        print("-------------------------------------------------")

    def ExportModel(self):
        headers = [
            ("STT:", self.stt),
            ("Mã bệnh nhân:", self.code),
            ("Tên bệnh nhân:", self.name),
            ("Ngày sinh:", self.datebirth),
            ("Tuổi:", self.age),
            ("Giới tính:", self.gender),
            ("Phường/Xã:", self.ward),
            ("Quận/Huyện:", self.district),
            ("Tỉnh/Thành Phố:", self.province),
            ("Ngày khám:", self.date),
            ("Ca Khám:", self.shift),
            ("ICD1:", self.primaryICD10Code),
            ("Tên ICD1:", self.primaryICD10),
            ("ICD2 (phụ):", self.secondICD10),
            ("Chuẩn đoán 2:", self.diagnosis),
            ("Tình trạng:", self.status)
        ]

        # Tính toán độ rộng cột
        max_key_length = max(len(header[0]) for header in headers) + 2
        max_value_length = max(len(str(header[1])) for header in headers) + 2

        s = ""
        s += "+" + "-" * max_key_length + "+" + "-" * max_value_length + "+\n"

        for header, value in headers:
            key_column = f" {header:<{max_key_length-2}} "
            value_column = f" {value:<{max_value_length-2}} "
            s += f"|{key_column}|{value_column}|\n"

        s += "+" + "-" * max_key_length + "+" + "-" * max_value_length + "+\n"
        return s

class DatePicker(tk.Toplevel):
    def __init__(self, parent, date_entry, next_widget):
        super().__init__(parent)
        self.title("Date Picker")
        self.overrideredirect(True)
        self.grab_set()
        self.date_entry = date_entry
        self.next_widget = next_widget

        self.cal = Calendar(self, selectmode="day", year=current_date.year, month=current_date.month,
                            day=current_date.day)
        self.cal.pack()

        confirm_button = customtkinter.CTkButton(self, text="Confirm", command=self.update_date)
        confirm_button.pack()

    def update_date(self):
        selected_date_str = self.cal.get_date()
        selected_date = datetime.datetime.strptime(selected_date_str, "%m/%d/%y")  # Convert to datetime object
        self.date_entry.delete(0, tk.END)
        self.date_entry.insert(0, selected_date.strftime("%d/%m/%Y"))
        self.destroy()
        self.next_widget.focus_set()
#function

def convert_date_format(date_str):
    try:
        date_obj = datetime.datetime.strptime(date_str, "%d/%m/%Y")
        return date_obj.strftime("%Y-%m-%d")
    except ValueError:
        print(f"Invalid date format: {date_str}")
        return date_str

def pick_date(event):
    next_widget = date_entry2
    date_picker = DatePicker(app, date_entry, next_widget)
    x, y, _, _ = event.widget.bbox("insert")
    x += event.widget.winfo_rootx() - 7
    y += event.widget.winfo_rooty() + 30
    date_picker.geometry("+%d+%d" % (x, y))

def pick_date2(event):
    next_widget = run_button
    date_picker = DatePicker(app, date_entry2, next_widget)
    x, y, _, _ = event.widget.bbox("insert")
    x += event.widget.winfo_rootx() - 7
    y += event.widget.winfo_rooty() + 30
    date_picker.geometry("+%d+%d" % (x, y))

def validate_dates():
    date_format = "%d/%m/%Y"
    date1_str = date_entry.get()
    date2_str = date_entry2.get()

    try:
        date1 = datetime.datetime.strptime(date1_str, date_format)
        date2 = datetime.datetime.strptime(date2_str, date_format)

        if date2 < date1:
            return False
        
        return True
    except ValueError:
        return False

def open_terminal_window():
    terminal_window = tk.Toplevel(app)
    terminal_window.title("Màn hình chạy dữ liệu")
    terminal_window.attributes("-fullscreen", True)
    icon_path = os.path.abspath("resource\crawlLogo.ico")
    terminal_window.iconbitmap(icon_path)
    terminal_text = tk.Text(terminal_window, bg="black", fg="green", insertbackground="green")
    terminal_text.pack(expand=True, fill='both')

    def on_closing():
        if script_thread.is_alive():
            terminal_window.destroy()
            app.deiconify()
        else:
            terminal_window.destroy()
            app.deiconify()

    terminal_window.protocol("WM_DELETE_WINDOW", on_closing)
    return terminal_window, terminal_text

def start_script_thread():
    global script_thread  # Khai báo biến toàn cục
    directory = filedialog.askdirectory(title="Chọn nơi lưu trữ!")
    if directory:
        directory = directory.replace('/', '\\')
    else:
        return
    terminal_window, terminal_text = open_terminal_window()
    app.withdraw()
    script_thread = threading.Thread(target=run_Script, args=(directory, terminal_text))
    script_thread.start()

# Hàm định dạng sheet
def center_window(window, width=600, height=440):
    # Lấy kích thước màn hình
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    # Tính toán tọa độ x và y để cửa sổ ở giữa màn hình
    x = (screen_width / 2) - (width / 2)
    y = (screen_height / 2) - (height / 2)
    window.geometry(f'{width}x{height}+{int(x)}+{int(y)}')

def tuple_to_dict(t):
    # Thay thế các tên cột này bằng tên thực tế của bạn
    columns = ['STT', 'Mã bệnh nhân', 'Tên bệnh nhân', 'Ngày sinh', 'Giới tính', 'Tuổi', 'Tỉnh Thành', 'Quận Huyện','Xã Phường', 'Ngày khám', 'Ca khám', 'Mã ICD chính','Tên ICD chính', 'Mã ICD phụ', 'Chuẩn đoán 2', 'Tình trạng']  
    return dict(zip(columns, t))
#app
def run_Script(directory,terminal_text):
    global  date_entry, date_entry2, app, current_date
    # Hàm để hiển thị thông tin lên terminal
    def log_terminal(message):
        terminal_text.insert(tk.END,message + "\n")
        terminal_text.see(tk.END)  # Scroll to the end
        terminal_text.update_idletasks()
    log_terminal(".........................Kiểm tra thời gian thực thi.....................................")

    if validate_dates() == True:
        log_terminal(".........................Thời gian hợp lệ................................................")
        start_date = datetime.datetime.strptime(date_entry.get(), "%d/%m/%Y")
        end_date = datetime.datetime.strptime(date_entry2.get(), "%d/%m/%Y")
        date_list = []
        current_date = start_date

        while current_date <= end_date:
            date_list.append(current_date)
            current_date += datetime.timedelta(days=1)
        
        log_terminal(".........................Xác định ngày thành công........................................")
        for d in date_list:
            mydb = connector.connect(
            host="192.168.0.127",
            port="3306",
            user="root",
            password="So17052001",
            database="nhidong2_full")
            mycursor = mydb.cursor()
            date = d.strftime('%Y-%m-%d')
            file_name = f"Patient_Data_{date}.xlsx"
            file_path = os.path.join(directory, file_name)
            log_terminal(".........................Tạo file Excel lưu thành công..................................")
            log_terminal(f".........................Nạp dữ liệu ngày {date}.......................................")
            query_string = f"CALL GetPatientReport('{date}')"
            listPatients = mycursor.execute(query_string, multi=True)
            all_results = []
            for result in listPatients:
                if result.with_rows:
                    all_results = result.fetchall()
                    break
            
            if len(all_results) > 0:
                patient_dict_list = [tuple_to_dict(pt) for pt in all_results]
                df = pd.DataFrame(patient_dict_list)
                # Lưu DataFrame vào Excel
                df.to_excel(file_path, index=False, sheet_name='Patient Data')
                
                # Mở workbook để điều chỉnh định dạng
                wb = load_workbook(file_path)
                ws = wb['Patient Data']
                
                # Tự động điều chỉnh chiều rộng cột
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Định dạng header
                header = ws[1]
                for cell in header:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

                # Bật tự động xuống dòng cho tất cả các ô
                for row in ws.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True)
                
                # Lưu lại workbook
                wb.save(file_path)
                log_terminal(f"...........................Nạp dữ liệu thành công vào file cho ngày {date}.........")
                log_terminal(f"...........................File đã được lưu: {file_path}...........................")
            else:
                log_terminal(f"...........................Không có dữ liệu cho ngày {date}........................")
            
            mydb.close()
        # Lưu file Excel
        
        log_terminal(".........................Hoàn thành xuất dữ liệu báo cáo..................................")
        app.deiconify()
    else:
        messagebox.showerror(title="Lỗi thời gian",message="Thời gian kết thúc phải lớn hơn hoặc bằng bắt đầu!")
        app.deiconify()

def run_secondary_interface(main_app):
    global run_button, date_entry, date_entry2, app, current_date
    app = customtkinter.CTkToplevel(main_app)
    app.title("Lấy dữ liệu khám có sẵn")
    center_window(app)
    icon_path = os.path.abspath("resource\crawlLogo.ico")
    app.iconbitmap(icon_path)

    def on_closing():
        main_app.deiconify()  # Hiển thị lại cửa sổ chính khi đóng cửa sổ mới
        app.destroy()

    app.protocol("WM_DELETE_WINDOW", on_closing)

    imgBG = ImageTk.PhotoImage(Image.open("resource\BG2.jpg"))
    l1 = customtkinter.CTkLabel(master=app, image=imgBG)
    l1.pack()

    frame = customtkinter.CTkFrame(master=l1, width=320, height=250, corner_radius=15)
    frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

    l2 = customtkinter.CTkLabel(master=frame, text="Select configuration", font=('Century Gothic', 20))
    l2.place(x=70, y=45)
    
    l3 = customtkinter.CTkLabel(master=frame, text="Từ ngày", font=('Roboto', 13))
    l3.place(x=20, y=90)
    date_entry = customtkinter.CTkEntry(master=frame, width=200, corner_radius=10)
    date_entry.place(x=100, y=90)
    date_entry.insert(0, current_date.strftime("%d/%m/%Y"))  # Đặt giá trị mặc định là ngày hiện tại
    date_entry.bind("<FocusIn>", pick_date)

    l4 = customtkinter.CTkLabel(master=frame, text="Đến ngày", font=('Roboto', 13))
    l4.place(x=20, y=130)
    date_entry2 = customtkinter.CTkEntry(master=frame, width=200, corner_radius=10)
    date_entry2.place(x=100, y=130)
    date_entry2.insert(0, current_date.strftime("%d/%m/%Y"))  # Đặt giá trị mặc định là ngày hiện tại
    date_entry2.bind("<FocusIn>", pick_date2)

    run_button = customtkinter.CTkButton(master=frame, command=start_script_thread,text="Thực thi", font=('Tahoma', 13), fg_color="#005369", hover_color="#008097")
    run_button.place(x=160, y=200)

    app.mainloop()