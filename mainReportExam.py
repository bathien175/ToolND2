#cào báo cáo dữ liệu khám
#using
import csv
import os
import time
import tkinter as tk
from tkinter import messagebox
from datetime import datetime, timedelta
import customtkinter
from tkinter import filedialog
from seleniumwire import webdriver
from selenium.webdriver.common.by import By
from PIL import ImageTk, Image
from tkcalendar import Calendar
from babel.numbers import *
import json
import pandas as pd
from tkinter import ttk
from openpyxl.cell import MergedCell
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
import threading

#global
customtkinter.set_appearance_mode("Light")
customtkinter.set_default_color_theme("green")
listStatus = []
current_date = datetime.datetime.today()
date_entry = Any
date_entry2 = Any
run_button = Any
app = Any

#class
class PatientModel:
    stt = 0
    code = ""
    dateExam = ""
    status = ""
    prescription = ""

    def to_dict(self):
        return {
            "STT": self.stt,
            "Mã bệnh nhân": self.code,
            "Ngày khám": self.dateExam,
            "Mã toa thuốc": self.prescription,
            "Tình trạng": self.status,
        }

    def printModel(self):
        print("-------------------------------------------------")
        print(f"STT: {self.stt}")
        print(f"Mã bệnh nhân: {self.code}")
        print(f"Ngày khám: {self.dateExam}")
        print(f"Mã toa thuốc: {self.prescription}")
        print(f"Tình trạng: {self.status}")
        print("-------------------------------------------------")

    def ExportModel(self):
        headers = [
            ("STT:", self.stt),
            ("Mã bệnh nhân:", self.code),
            ("Ngày khám:", self.dateExam),
            ("Mã toa thuốc:", self.prescription),
            ("Tình trạng:", self.status)
        ]

        # Đảm bảo tất cả các giá trị không phải là None
        headers = [(header, "" if value is None else value) for header, value in headers]

        # Tính toán độ rộng cột
        max_key_length = max(len(header[0]) for header in headers) + 2
        max_value_length = max(len(str(header[1])) for header in headers) + 2

        s = ""
        s += "+" + "-" * max_key_length + "+" + "-" * max_value_length + "+\n"

        for header, value in headers:
            key_column = f" {header:<{max_key_length-2}} "
            value_column = f" {value:<{max_value_length-2}} "
            s += f"|{key_column}|{value_column}|\n"

        s += "+" + "-" * max_key_length + "+" + "-" * max_value_length + "+\n"
        return s

class StatusExamModel:
    # thuộc tính
    status_code = 0
    status_name = ""
    # phương thức
    def setName(self, status_name):
        self.status_name = status_name
    
    def getName(self):
        return self.status_name
    
    def setCode(self, status_code):
        self.status_code = status_code
    
    def getCode(self):
        return self.status_code
    
class DatePicker(tk.Toplevel):
    def __init__(self, parent, date_entry, next_widget):
        super().__init__(parent)
        self.title("Date Picker")
        self.overrideredirect(True)
        self.grab_set()
        self.date_entry = date_entry
        self.next_widget = next_widget

        self.cal = Calendar(self, selectmode="day", year=current_date.year, month=current_date.month,
                            day=current_date.day)
        self.cal.pack()

        confirm_button = customtkinter.CTkButton(self, text="Confirm", command=self.update_date)
        confirm_button.pack()

    def update_date(self):
        selected_date_str = self.cal.get_date()
        selected_date = datetime.datetime.strptime(selected_date_str, "%m/%d/%y")  # Convert to datetime object
        self.date_entry.delete(0, tk.END)
        self.date_entry.insert(0, selected_date.strftime("%d/%m/%Y"))
        self.destroy()
        self.next_widget.focus_set()

def loadingData():
    s1 = StatusExamModel()
    s2 = StatusExamModel()
    s3 = StatusExamModel()
    s4 = StatusExamModel()
    s1.setCode(0)
    s1.setName("Chưa khám")
    s2.setCode(1)
    s2.setName("Đã khám")
    s3.setCode(3)
    s3.setName("Chờ CLS")
    s4.setCode(5)
    s4.setName("")
    listStatus.append(s1)
    listStatus.append(s2)
    listStatus.append(s3)
    listStatus.append(s4)

def convert_date_format(date_str):
    try:
        date_obj = datetime.datetime.strptime(date_str, "%d/%m/%Y")
        return date_obj.strftime("%Y-%m-%d")
    except ValueError:
        print(f"Invalid date format: {date_str}")
        return date_str

def pick_date(event):
    next_widget = date_entry2
    date_picker = DatePicker(app, date_entry, next_widget)
    x, y, _, _ = event.widget.bbox("insert")
    x += event.widget.winfo_rootx() - 7
    y += event.widget.winfo_rooty() + 30
    date_picker.geometry("+%d+%d" % (x, y))

def pick_date2(event):
    next_widget = run_button
    date_picker = DatePicker(app, date_entry2, next_widget)
    x, y, _, _ = event.widget.bbox("insert")
    x += event.widget.winfo_rootx() - 7
    y += event.widget.winfo_rooty() + 30
    date_picker.geometry("+%d+%d" % (x, y))

def validate_dates():
    date_format = "%d/%m/%Y"
    date1_str = date_entry.get()
    date2_str = date_entry2.get()

    try:
        date1 = datetime.datetime.strptime(date1_str, date_format)
        date2 = datetime.datetime.strptime(date2_str, date_format)

        if date2 < date1:
            return False
        
        return True
    except ValueError:
        return False

def open_terminal_window():
    terminal_window = tk.Toplevel(app)
    terminal_window.title("Màn hình chạy dữ liệu")
    terminal_window.attributes("-fullscreen", True)
    icon_path = os.path.abspath("resource\crawlLogo.ico")
    terminal_window.iconbitmap(icon_path)
    terminal_text = tk.Text(terminal_window, bg="black", fg="green", insertbackground="green")
    terminal_text.pack(expand=True, fill='both')

    def on_closing():
        if script_thread.is_alive():
            terminal_window.destroy()
            app.deiconify()
        else:
            terminal_window.destroy()
            app.deiconify()

    terminal_window.protocol("WM_DELETE_WINDOW", on_closing)
    return terminal_window, terminal_text

def start_script_thread():
    global script_thread  # Khai báo biến toàn cục
    directory = filedialog.askdirectory(title="Chọn nơi lưu trữ!")
    if directory:
        directory = directory.replace('/', '\\')
    else:
        return
    terminal_window, terminal_text = open_terminal_window()
    app.withdraw()
    script_thread = threading.Thread(target=run_Script, args=(directory, terminal_text))
    script_thread.start()
# Hàm định dạng sheet
def center_window(window, width=600, height=440):
    # Lấy kích thước màn hình
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    # Tính toán tọa độ x và y để cửa sổ ở giữa màn hình
    x = (screen_width / 2) - (width / 2)
    y = (screen_height / 2) - (height / 2)
    window.geometry(f'{width}x{height}+{int(x)}+{int(y)}')
    
def format_sheet(ws, date_str):
    # Dòng tiêu đề chính
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = "Báo cáo sổ khám bệnh"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal="center")

    # Dòng ngày khám
    ws.merge_cells('A2:E2')
    date_cell = ws['A2']
    date_cell.value = f"Ngày khám: {date_str}"
    date_cell.font = Font(size=14, bold=True)
    date_cell.alignment = Alignment(horizontal="center")

    # Dòng header các cột
    headers = ["STT", "Mã bệnh nhân", "Ngày khám", "Mã toa thuốc", "Tình trạng"]
    header_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = Font(size=13, bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = header_fill

    # Tự động điều chỉnh độ rộng cột
    for col in ws.iter_cols(min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        max_length = 0
        column = col[0].column_letter  # Lấy tên cột
        for cell in col:
            if isinstance(cell, MergedCell):
                continue  # Bỏ qua ô đã hợp nhất
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

def find_status(id):
    for s in listStatus:
        if s.getCode() == id:
            return s.getName()
    return None

def run_Script(directory,terminal_text):
    global  date_entry, date_entry2, app, current_date
    # Hàm để hiển thị thông tin lên terminal
    def log_terminal(message):
        terminal_text.insert(tk.END,message + "\n")
        terminal_text.see(tk.END)  # Scroll to the end
        terminal_text.update_idletasks()
    ascii_art = """
    ---::::.:-=-----=----:--=+++++=-==----=++**+===+++*#**+===---------------------=#%%%%%%*-:::::::::::::::-=-:-++=-:-+*+-::=*+-:::-::--::-=%
    -::::::.:-=----------::-=+++++=--::::::-=++=======++++==---------::::::::------=#%%%%%%*-:::::::::::::::-=-:-===-:-=+=-:-+#+-::=*=::::-::%
    -::::::.:-=----------::-=+**++=-:::--:::-===-=+===+====---------::.......::--:--*%%%%%%#-:::::::::::::::---:-===-:-===---+*=-:-*#+-:-++-:#
    ::::::..:-=-:--------::-=++*++=-::-==-:::-====*+====-==-------::..:----:..::-::-*%%%%%##-:::::::::::::::--::-==--::-==-:-=+=-:-+*=-:-#*=:#
    ::::::.::---:--------::-=++*++=-:------::-==-==+====-=--------:..:--===-:..:::--*%%%%%##=:::::::::::::::--:::-=--:--=--:--==-:-=+=--=*+-:#
    :::..::::---:--------::-=+**++=-:-----::::-=-======--=-------::.::-------:.::::-+%%%%%##+:::::::::::::::--:::-=--::-=--:-==--::===---+==-%
    ::::::::----:--------:::=+**++=-::----::::----=====--=-------:...:::-:--::.::---=#%%%%##+-::::::::::::::--:::----::-=--::-=--::==--:-===-%
    ---::::::---::-------::-=+++++=-::----::::-----=====---------:...:::-::::..:::--=#%%%%%#+-::::::::::::::--:::----::----::-=--::-=----==--%
    ----:::::::::::------::-=+++++=-::::::::::------=-==---------:....::---::..::::-=#%%%%%#*-::::::::::::::--:::----::----::----::-=--:-==--%
    ::::::::::::::::::::::::-++++++--::::::::---------==---------::...::::::..::::::=#%%%%%#*-::::::::::::::--:::----::----:::---::----:-=---%
    :::::::::::::::::::--:::-++*+++---::::::----------------------::.........::::::-=#%%%%%#*-::::::::::::::--:::----:::----:----::----:-=---%
    ::::::::::::::::::::::::=++**++=-=-----------------------------:::.....:::::::::=#%%%%##*::...::::.....::-:::----:::---::----::----:-----%
    ::::::::::::::::::::::::-++**++--======------------------::----:::::::::::::::::-#%%%%%#*:..............::-:-----:::---::----::----:-----#
    ::::::::::::::::::::::::-=+**++-::----===-::::::---------:::-=---=-:::....::::::=#%%%%%%*-:::::::::::::::::::::---:----::----::----:-----#
    ::::::::::::::::::::::::-=+***+-:::::::---:::::::::----:::-=-:--=+****+==-------+%%%%%%%*-::::::------------:::::::-----:----::----:-----%
    ===---::::::::::::::::::-=+***+-:::::::::::::::::---=--:-=====+++***######*+=---+%%%%%%%*-:::::::::::::--::::--=------:::::--------------#
    ****+++====----:::::::::-=+****-:::::::::::::::::::-----=:::=++*#*#############**%%%%%%%*-:::::::::::::::::::::::::-=====---::::---------%
    ************++++===----:-+****+-::::::::::::::::::----=***==*#######%%%%%%%%%%%%%%%%%%%%+::::::::::::::::::::::::::::::--=+=--=====------#
    *****+++*************++==+*****=::::::::::::::::----==***#####%%%%%%%%%%%%%%%%%%%%%%%%%%+:::::-:::::::::::::::::::::::::::-:::-========--#
    ++++++++**********++***=-=+****++===----:::----=+***#######%%%%%%%%%%%%%%%%%%%%%%%%%%###+::::::::::::::::--::::::::::--:::::::::::::-----%
    --=======++++++++******--=+*********+++++++===+#############%%%%%%%%%%%%%%%%%%%%%%%%%%%%%*=::::::::::::::::::-::::::::::::------:::::::::#
    ------------======++**+--=**#****+++++++*****###%#########%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%*-:::::::::::::::::::::::::::::::::::-------::#
    -----=-------==---=====--+*##***++++++****###%%%%%%%%###%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%*-::::::::::::::::::::::::::::::::::::-------#
    ------------------====--=+*##***=====+***#####%%%%%#####%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%*::::::::::::::::::::::::::::::::::---------#
    --------------------==--=+*##***=--==+*+*###%#%%%%####%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%=:::::::::::::::.::::::::::::::::::::::::::#
    ::::----------------=-:--+*##**+=---=**#%##%%%%%%%##%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%::::::::::::::::...:::::::::::::::::::::::#
    :::.::-::::-----------::-=*##*++----+*#%%%%%%%%%%%%%%#%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%+:::::::::::::::::::::..::.:::::::::::::::#
    -++-:.:::::::::::::---::-+*###**----+##%###%%%%%%%######%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%=:::::::::::::::::::::::::::::::::::::::::#
    +##+-:::-:::::::::::-:::-=***#**-:-=*=#%##%%%%%%#******####%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%+::..:::::.:::::::::::::::::::::::::::::::#
    +**+-:.:-::::::--:::-:::-=*****+-::-*+#%%%%%%%#*++==++++**#####%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%*:::::::::::..::::::::::::::::::::::::::::#
    +**=::.::::::::------:::--=+++++=::=##%%%%%%%#+========+++***#####%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%:::::::::::::::::::::::::::::::::::::::::#
    +++=:...:::::::------:::-+******+-:=#%%%%%%%%*==----=====++++****#####%%%%%%%%%%%%%%%%%%%%%%%%%%+:::::::::::::::::::::::::::::::::::.::::#
    +++-:...:::::::--:::::::-++++**++--+#%%%%%%%%+==--------======+++****##%%%%%%%%%%%%%%%%%%%%%%%%%%-:::::::::::-:::::::::::::::::::::::::::#
    =++-:...:::::::--:::::::-+++**+++--*#%%%%%%%#+==--------------====+++++**#%%%%%%%%%%%%%%#%%%%%%%%+::::::::::----===-:::::::::::::::::::::#
    =++-:..::::::::--:::::::-+######+-:+#%%%%%%%#+=-----------=++*****+=====++*#%%%%%%%%%%**+==#%%%%%%-::::::::------===::::::::::-----::::::#
    +++-:...::::::---::::::::=++++++=::-#%%%%%%%#+-------==++#####*+++++==---===++=+#%%%#+#****=%%%%%%+::::::::==-::-----:::::::::----===::::#
    +*+=:...::::::---:::::::-==++++++::-+#%%%%%%#+========**##******++====---:::-+#%%%#*+++==-++#%%%%%%+:::::::=+-::::::-::::::::--::--==-:::#
    +++=:...::::::---:::::::=**###***=:-++##%##%%*###**++++*+*%%%%#*#%#==-:::-==--+###=-=+##*+-+*%%%%%%%#-:::::-=-::::::-:.::::::=-:::----:::#
    +++-:...:::::::--:::::::-+######+::-+++=+%*%%#####**++*#%%%%%%%%%%%#=------:::-+*+-:-=*++*==*%%%%%%%%%+:::----::::---::::::::=-::::---:::#
    ++=-:...::::::---::::::::-++***+=:.:+++-:-###*+++*###**%%%%%%%%%%%##*----::::::----:::++===-#%%%%%%%%%%+-:---:::::-==----------::::--::::#
    +==-:..:::::::----:::::::-+++*++=:.:++=-:.:*##%%%#***++%%%%%%%%%%%##*+-:::::::::---::-=-::-=%%%%%%%%%%%#-:::::::::-==---:-----:::::------#
    +++=:...::::::-------::::-=++++++-::=+=-:.+%%%%%%%%*+*=*####%%%%%##*++-:::::::-----::---::-#%%%%%%%%%%%%+++++=:..:-==--::::::::::::-==---#
    ----::..:::::::------:::-=+++++++=:-+++-::#%%#%%%%%%++--+########**++=-::::--------:::----=+#%%%%%%%%%%%%%%%###+-:-----::::::::::::-==--:#
    ::::::::::::::::::::::::-+***+***+::-==-:.#%%##%%%%%*+::-=+****+++++==-::::---------:-+=+===+*%%%%%%%%%%%%%%%%#####*+--::::::::..::-----:#
    -------:--::::--::-:::::-+**####*+::::::.:+#%#*#%%%%*=:::-==+++++====-----------------=+==--==#%%%%%%%%%%%%%##%########*=:........:::::::#
    :------------::::---------===+++=-::::--:::*#######**-:::::-=+++++==-------------------====--=*%%%%%%%%%%%%%%%############*=:::...:::::::#
    ----::::::::::::::::::::---=======-::=+=--:-++*+***++-::::::-=+===--------------------========#%%%%%%%%%%%%%%%%%#############+=-:.:::::::#
    *+---:::::::::::::::::::=++++++++++=+*+===-::-+++++++=+++*#*++=-==-------------------=========%%%%%%%%%%%%%%%%%################*-:::::::.#
    #+--:::::::-::::::.....:=*****#****+*####*=:..:-=+++++###*++==------===-----------====++++===*%%%%%%%%%%%%%%%%%%%##%%%%%%%#######+:::::..#
    *=::::::---==-:::::....:-*########*+++***-:.....:=====+++==-------=====-----==-=====++*+++==+#%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%#######=::::.#
    *+---------===---::.....:-+++++++==------::......:=======--:::--======-----========+****++==*#%%%%%%%%%%%%%%%%%%%%%%%%%%%%##########*:...#
    #+=-------=====--::....::---======--:::-==-:......-=======+*******++=============+**##*++==+*%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%#######=..#
    #=--------=====--:::..::----=======----=*+=-::.....-=+++####**=================++*###**+===+*%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%##%%#+.#
    #=---------====--::::::====++*+++++++==+**+=-::....:-++*#+====++=============++*#####*++==++#%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%#%%%*%
    #+----------===--:---::+*###*####****=+***+=--::----+*+++++++****++========++*######**++=+++#%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%#%%%
    #+=--------=---------:-=***#**###****+*#**++=---**++*##*++**###**+==---==++*##%####***+++++*%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%#%
    #=-:::-----------=+=..:-++++++*++++=++**####*=--=+*###%%*++***+==---====+*##%%#####****++++*%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
    #=-:..:-------===++=...:=++++++***+++=*##*#%%##*+*%%%%%%%%++=======+++**##%#%#######****+++#%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
    #=-::::-------=--=++-:::=++++++++++++++****#%####%%%%%%%%%%*++++***###%%%###########*******#%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
    #=-::::-----------=++======++++++++++++++#######%%%%%%%%%%%%%####%%%%%%%%%###########****##%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
    #=----==========---=+*##***********+**++=*%###%%%%%%%%%%%%%%%%%%%%%%%%%%%%%###########****%#%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
    #+=*#%%*=============+**+++++++***+++++=+*%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%##########***%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
    +=-+**+*+==-----=====++==+++++*****+*****#%%%#%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%########***%#%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
    -------=###=-::-====-:::-+++++*****+*##%%%%%%#%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%#######**%%#%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
    -------=###=-:-===-...:-=++*++++**++*%%######%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%#####***%%%#%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
    =-------*##=-:-==-...-+*++++++++++++*#####%#%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%#****#%%%%#%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
    --------*##+===**=:-+##*+++++++++++++*######%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%##*#%%%%%#%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
    ====--==*##+=*%%%%#%%%%#+++++++*******####%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%#%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
    ++++++++*##*=+*######%%#+++++++*+++++*###%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%#%%%%%%%%%%%%%%#%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
    -------=*##+---==++*##%#+++++++****++*###%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%#%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
    -------=*##+---==+*#%%%#+++++++**++==+###%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%#%%%%%%%%%%%%%#%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
    -------=*##+---+###%%#+++++++++**+===+###%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%#%%%%%%%%%%%%%#%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
    -------=*##+--=*####%*++++=++++**+===+###%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%#%%%%%%%%%%%%%#%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
    -------=*##+-=+#***##++++==++++**+===+%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%#%%%%%%%%%%%#%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
    -=======*##+=+*#*=*#+=+++==++++**+===+%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%*%%%%%%%%%%%#%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
    ======-=*##+=*##*+**=+*+==+++++++++=+#%%%%%%%%%%%%%%%%%%%%%%%%%%%%##%%%%%%%##%%%%%%%%%#%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
    ========*##++##*+**+=**+===+++++++++#%%%%%%%%%%%%%%%%%%%%%%%%%%%*-::-*%%%%%%+%%%%%%%%%#%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
    ========*##**######***+======++++++#%%%%%%%%%%%%%%%%%%%%%%%%%%*-:-::::=**%%%##%%%%%%%*###%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
    =======+***+*##+*##**++========+++#%%%%%%%%%%%%%%%%%%%%%%%%#=:--===-:::---===-###%%%+*###%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
    =======+**++*%#*==+++==========++*%%%%%%%%%%%%%%%%%%%%%%**+:::-==+*+-:::-::::-+##*#***####%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
    ------------+%%+-===-======+====+#%%%%%%%%%%%%%%%%%%%%#-::::::-=+##%+=------:::++*+*++#%%##%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
    ------------=#%#+++-===========+#%%%%%%%%%%%%%%%%%%%%%+::::::::::-=**+==---:::-*+=+***%%%%#%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
    -------------+###*=-=+++==++==+#%%%%%%%%%%%%%%%%%%%%%#:::::--:::::::-**=--::::=+++*#*#%%%%#%%%%%%%%%%%%%%%%%%%%%#%%%%%%%%%%%%%%%%%%%%%%%%%
    --------------=======+++++++++#%%%%%%%%%%%%%%%%%%%%%%=:::::-===-----=+##*++*---=#*+#%%%%%%%%###%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
    --------------------=++++++*+-#%%%%%%%%%%%%%%%%%%%%%+::::::::=***+====###=-=--=#+*#*#%%%%%#%%%%##%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
    --------------------=+*****+-:-+#%%%%%%%%%%%%%%%%%%*:::::::::::=*#====+#------*%%++#*#%%%###**###%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%##*%
    ---------------------+*****-:::--#%%%%%%%%%%%%%%%%%:::::-=--:::::=+===+*#*----#%%++*%*#%%###**###%%%%%%%%%%%%%#**##%%%%%%%%%%%%%%%%#*+=--#
    """
    log_terminal(ascii_art)
    log_terminal("......................................................................................................................................")
    log_terminal(".........................Kiểm tra thời gian thực thi.....................................")

    if validate_dates() == True:
        log_terminal(".........................Thời gian hợp lệ................................................")
        start_date = datetime.datetime.strptime(date_entry.get(), "%d/%m/%Y")
        end_date = datetime.datetime.strptime(date_entry2.get(), "%d/%m/%Y")
        date_list = []
        current_date = start_date

        while current_date <= end_date:
            date_list.append(current_date)
            current_date += datetime.timedelta(days=1)
        
        log_terminal(".........................Xác định ngày thành công........................................")
        file_name = f"Dữ liệu khám {start_date.strftime('%d-%m-%Y')}_{end_date.strftime('%d-%m-%Y')}.xlsx"
        file_path = os.path.join(directory, file_name)
        # Tạo hoặc mở file Excel
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
        else:
            wb = Workbook()
            wb.remove(wb.active)  # Xóa sheet mặc định khi tạo mới
        log_terminal(".........................Tạo file Excel lưu thành công...................................")
        log_terminal(".........................Khởi động Chrome................................................")
        #chrome settings
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_experimental_option("prefs", {
            "download.default_directory": directory,
            "profile.default_content_setting_values.automatic_downloads": 1,
            "download.prompt_for_download": False,
            "profile.default_content_settings.popups": 0,
            "safebrowsing.enabled": "false",
            "safebrowsing.disable_download_protection": True
        })
        prefs = {"credentials_enable_service": False,
            "profile.password_manager_enabled": False}
        chrome_options.add_experimental_option("prefs", prefs)
        log_terminal(".........................Đang cấu hình chrome............................................")
        chrome_options.add_argument("--headless")  # Chạy trình duyệt trong chế độ headless
        chrome_options.add_argument("--disable-gpu")  # Tăng tốc độ trên các hệ điều hành không có GPU
        chrome_options.add_argument("--window-size=1920x1080")  # Thiết lập kích thước cửa sổ mặc định
        driver = webdriver.Chrome(options=chrome_options)
        log_terminal(".........................Cấu hình thành công.............................................")
        url = "http://192.168.0.77/dist/#!/login"
        driver.get(url)
        time.sleep(3)
        log_terminal(".........................Duyệt WEBSITE thành công........................................")
        # login
        form = driver.find_element(By.NAME, "LoginForm")
        username = driver.find_element(By.NAME, 'username')
        password = driver.find_element(By.NAME, 'password')
        username.send_keys('quyen.ngoq')
        password.send_keys('74777477')
        # ấn nút login
        form.submit()
        time.sleep(2)
        log_terminal(".........................Đăng nhập ẩn thành công.........................................")
        log_terminal(".........................Hack vào WEBSITE thành công.....................................")
        ascii_nd2 = """
    .................................................................  ..-=+**++=====+**+=-:. .................................................................. 
    .................................................................:-+*+=-....:..:.:....:-+*+=:.................................................................
    .............................................................:=**+-..+=..*-+. ==+.+:=.-..:=**=:..............................................................
    ...........................................................:+**-..:-:==-.=.-..:.-.-:-=:..=:.-+*+-............................................................
    .........................................................:+**-..:-+.=:--:--======--:::..+-+..:=+**-..........................................................
    ........................................................=**+..-...-.:-===============+=::-.=-+:.=**+:........................................................
    ......................................................:***-.:-==..-=====================+-.---.-.:***-.......................................................
    .....................................................-***:.----.:==========================-.:=+-:.+**=......................................................
    ....................................................-***:.-+=-.-=+*+=======+****+========*+==..+:-..***=.....................................................
    ...................................................:***-.:.:-.===+*+======********+=====+#*===..+-=.:***-....................................................
    ...................................................+**+:====.====+**=====**********=====+#+====.:-...=***....................................................
    ..................................................-***:.:--::+====**=====**********=====**+====-...:..***=...................................................
    ..................................................+***.-++=.+=====+**=====********+====+#*=====+.:+-=.=***...................................................
    ..................................................***=.....:+======***+====+****+====+***=======-.....:***:..................................................
    .................................................:***-.....-========+*****+======+******=========......***-..................................................
    .................................................:***-.....===========+****======****++=========+======***=..................................................
    .................................................:***-.....=====================================+*********-..................................................
    ..................................................***=.....=+++++++++++++==========++++++++++++++*********:..................................................
    ..................................................+***------------------=++++==+++==------------------+***...................................................
    ..................................................:***::::::::::::::::::-+***==***+=::::::::::::::::::***=...................................................
    ...................................................+**+-----------------==*##+=**#+=-----------------+***....................................................
    ....................................................***-::::::::::::::::==*##+=**#+=::::::::::::::::-***:....................................................
    ....................................................:***=---------------+=***+=***++----------------***-.....................................................
    .....................................................:***-::::::::::::::+=***+=**#++::::::::::::::-***-......................................................
    .......................................................+**=-------------+=***==***++=------------=***:.......................................................
    ........................................................-**+--------:::===***==***+==::---------+**=.........................................................
    ..........................................................=**=---------+==***==***==+---------=**+:..........................................................
    ............................................................=**+-------+==***==***==+=------=**=:............................................................
    ..............................................................-+*+=---====***==***====----=*+-...............................................................
    .................................................................:=+++*++=+++==+*+=++++++=-..................................................................
    .....................................................................:--=++++++++++==-:......................................................................
    .............................................................................................................................................................
    .............................................................................................................................................................
    .............................................................................................................................................................
    .............................................................................................................................................................
    ..............................................................................................................................................................
    .......:......................................................................................................................................................
    :===*+-+++.................:==.....................++-..............:==........:==..............................**=...............:-:.........................
    ....*..*:*..:+:..........-.--=..=..=-..............+-=....-.....:...=:=..-:....=:=...............:+:.:+.........#-+..==.......:...-=-..........#..............
    ..:*..*==.:+=+.---....:-+:=-=..-.:--.:-::....-:-.:++:.+.=+....--=:.=--.::.-.:.=--.....=:-.==-+::+=+.:-....---::**:.-++=.....+=+.=-:+.:===.....#+=+-:+==.---:.
    ..:+..**+.*..*:==#:....--.=++.:=.+-=.:=+-....==*:-**-:+.=-....=.::.===.--.==-.===....:*++:*.=*+*..:+==....=+%-:##=-*.:#....=-:=.+::*.*:.+:...:#-.-*#:.*:=+#-.
    ..-+..*=*-*.:*:=*+=....--.===:-=-=:.-:+==....=*=--*-=:+.*=....=..::=-=:--.+-=.=-=:....#+*=*=-*=*:.==-=....=#+=-%=*++.=#....==..=+-=*-*.:*-...:#..*=#.:#:=#+=.
    ..:=..=::--==:-:-.-.....-:::.-:-.-=-:.-.::...::.-:=.---=:-.....---.-::-:-.-.:::::-....=.---=-+-:==-..=....:-.=:+.=-=+--:....-==::=-:::==:-....===:.:++:-:-.=.
    ..................................:-..........................................................*-.....................................................--.......
    .............................................................................................=:..................................................... ........
                """
        log_terminal(ascii_nd2)
        # Chọn
        liElement = driver.find_element(By.XPATH,
                                        "//li[contains(@class, 'nav-item') and contains(@permission, "
                                        "'report_menu_all')]")
        liElement.click()
        time.sleep(1)
        log_terminal(".........................Chuyển vào xuất report..........................................")
        sokham = driver.find_element(By.XPATH, "//a[contains(text(), 'Sổ Khám Bệnh')]")
        sokham.click()
        time.sleep(1)
        log_terminal(".........................Chọn sổ khám bệnh...............................................")
        for date in date_list:
            # Điền thông tin
            dateinput = driver.find_element(By.XPATH, "//input[contains(@id, 'inputjqxDateTimeInput') and @type='textarea']")
            driver.execute_script("arguments[0].value = arguments[1]", dateinput, date.strftime("%d/%m/%Y"))
            driver.execute_script("arguments[0].dispatchEvent(new Event('input', { bubbles: true }))", dateinput)
            time.sleep(0.5)
            log_terminal(".........................Chọn ngày cần xuất báo cáo......................................")
            inputZone = driver.find_element(By.XPATH, "//div[@id='dropdownlistContentzone_exam']//input[@class='jqx-combobox-input jqx-combobox-input-energyblue jqx-widget-content jqx-widget-content-energyblue jqx-rc-all jqx-rc-all-energyblue']")
            inputZone.clear()
            inputZone.send_keys('Nguyễn Du')
            time.sleep(0.5)
            log_terminal(".........................Chọn khu làm việc...............................................")
            inputShift = driver.find_element(By.XPATH, "//div[@id='dropdownlistContentworkInShift']//input[@class='jqx-combobox-input jqx-combobox-input-energyblue jqx-widget-content jqx-widget-content-energyblue jqx-rc-all jqx-rc-all-energyblue']")
            inputShift.clear()
            inputShift.send_keys('Tất cả')
            time.sleep(0.5)
            log_terminal(".........................Chọn ca làm việc ...............................................")
            btnSearch = driver.find_element(By.ID,'btXemBaoCao')
            btnSearch.click()
            for i in range(1,11):
                time.sleep(1)
                log_terminal(f".........................Đang tải dữ liệu báo cáo {i*10}%....................................")
            log_terminal(".........................Tải dữ liệu thành công..........................................")
            demstt = 1
            listPatient = []
            for request in reversed(driver.requests):
                if request.response:
                    if f'out_patient_new_regist/baoCaoSoKhamBenh?date={date.strftime("%Y-%m-%d")}&zoneCode=nguyendu&shift=-1' in request.url:
                        # format cho dữ liệu về dạng json
                        response_data = json.loads(request.response.body)
                        # đọc data
                        patient_info = response_data.get('data', {})
                        for pi in patient_info:
                            patientmd = PatientModel()
                            patientmd.stt = demstt
                            patientmd.code = pi.get('patient_code')
                            patientmd.dateExam = date.strftime("%Y-%m-%d")
                            patientmd.status = find_status(pi.get('status'))
                            patientmd.prescription = pi.get('prescription_id')
                            listPatient.append(patientmd)
                            log_terminal(patientmd.ExportModel())
                            demstt+=1
                        break
            # Chuyển listPatient thành DataFrame
            patient_dict_list = [pt.to_dict() for pt in listPatient]
            df = pd.DataFrame(patient_dict_list)
            log_terminal(".........................Chuyển dữ liệu thành công.......................................")
            # Tạo sheet mới cho từng ngày
            sheet_name = date.strftime("%d-%m-%Y")
            ws = wb.create_sheet(title=sheet_name)
            log_terminal(".........................Tạo Sheet thành công............................................")
            # Thêm dữ liệu vào sheet
            for r in range(3):
                ws.append([])

            for r in dataframe_to_rows(df, index=False, header=False):
                ws.append(r)

            # Định dạng sheet
            format_sheet(ws, sheet_name)
            log_terminal(".........................Định dạng sheet thành công......................................")
        # Lưu file Excel
        wb.save(file_path)
        log_terminal(".........................Hoàn thành xuất dữ liệu báo cáo.................................")
        app.deiconify()
    else:
        messagebox.showerror(title="Lỗi thời gian",message="Thời gian kết thúc phải lớn hơn hoặc bằng bắt đầu!")
        app.deiconify()

#app
def run_secondary_interface(main_app):
    global run_button, date_entry, date_entry2, app, current_date
    app = customtkinter.CTkToplevel(main_app)
    app.title("Lấy báo cáo sổ khám bệnh")
    center_window(app)
    icon_path = os.path.abspath("resource\crawlLogo.ico")
    app.iconbitmap(icon_path)

    def on_closing():
        main_app.deiconify()  # Hiển thị lại cửa sổ chính khi đóng cửa sổ mới
        app.destroy()

    app.protocol("WM_DELETE_WINDOW", on_closing)

    imgBG = ImageTk.PhotoImage(Image.open("resource\BG2.jpg"))
    l1 = customtkinter.CTkLabel(master=app, image=imgBG)
    l1.pack()

    frame = customtkinter.CTkFrame(master=l1, width=320, height=250, corner_radius=15)
    frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

    l2 = customtkinter.CTkLabel(master=frame, text="Select configuration", font=('Century Gothic', 20))
    l2.place(x=70, y=45)

    loadingData()
    
    l3 = customtkinter.CTkLabel(master=frame, text="Từ ngày", font=('Roboto', 13))
    l3.place(x=20, y=90)
    date_entry = customtkinter.CTkEntry(master=frame, width=200, corner_radius=10)
    date_entry.place(x=100, y=90)
    date_entry.insert(0, current_date.strftime("%d/%m/%Y"))  # Đặt giá trị mặc định là ngày hiện tại
    date_entry.bind("<FocusIn>", pick_date)

    l4 = customtkinter.CTkLabel(master=frame, text="Đến ngày", font=('Roboto', 13))
    l4.place(x=20, y=130)
    date_entry2 = customtkinter.CTkEntry(master=frame, width=200, corner_radius=10)
    date_entry2.place(x=100, y=130)
    date_entry2.insert(0, current_date.strftime("%d/%m/%Y"))  # Đặt giá trị mặc định là ngày hiện tại
    date_entry2.bind("<FocusIn>", pick_date2)

    run_button = customtkinter.CTkButton(master=frame, command=start_script_thread,text="Thực thi", font=('Tahoma', 13), fg_color="#005369", hover_color="#008097")
    run_button.place(x=160, y=200)

    app.mainloop()