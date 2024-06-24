#cào toa thuốc
#using
import csv
import os
import time
import tkinter as tk
from tkinter import messagebox
from datetime import datetime, timedelta
import customtkinter
from tkinter import filedialog
from seleniumwire import webdriver
from selenium.webdriver.common.by import By
from PIL import ImageTk, Image
from tkcalendar import Calendar
from babel.numbers import *
import json
import pandas as pd
from tkinter import ttk
from openpyxl.cell import MergedCell
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
import threading
#global
customtkinter.set_appearance_mode("Light")
customtkinter.set_default_color_theme("green")
pathExcel = ""
#class

#function
def center_window(window, width=600, height=440):
    # Lấy kích thước màn hình
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    # Tính toán tọa độ x và y để cửa sổ ở giữa màn hình
    x = (screen_width / 2) - (width / 2)
    y = (screen_height / 2) - (height / 2)
    window.geometry(f'{width}x{height}+{int(x)}+{int(y)}')
#app
def run_secondary_interface(main_app):
    global run_button, date_entry, date_entry2, app, current_date
    app = customtkinter.CTkToplevel(main_app)
    app.title("Lấy dữ liệu toa thuốc")
    center_window(app)
    icon_path = os.path.abspath("crawlLogo.ico")
    app.iconbitmap(icon_path)

    def on_closing():
        main_app.deiconify()  # Hiển thị lại cửa sổ chính khi đóng cửa sổ mới
        app.destroy()

    app.protocol("WM_DELETE_WINDOW", on_closing)

    imgBG = ImageTk.PhotoImage(Image.open("BG.png"))
    l1 = customtkinter.CTkLabel(master=app, image=imgBG)
    l1.pack()

    frame = customtkinter.CTkFrame(master=l1, width=320, height=300, corner_radius=15)
    frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

    l2 = customtkinter.CTkLabel(master=frame, text="Select configuration", font=('Century Gothic', 20))
    l2.place(x=70, y=45)

    drop = customtkinter.CTkButton(master=frame, command=select_Excel_file, width=300, height=50, corner_radius=15, fg_color="#FF9292", hover_color="#FFB6B6", text_color="#000000", text="Ấn để chọn file Excel!", font=('Tahoma', 13))
    drop.place(x=10, y=100)

    cancel_button = customtkinter.CTkButton(master=frame, text="Hủy bỏ", command=cancel_Excel_file, font=('Tahoma', 13), fg_color="#B3271C", hover_color="#FF3D4D")
    cancel_button.place(x=10, y=200)

    run_button = customtkinter.CTkButton(master=frame, text="Thực thi", command=handle_file_excel, font=('Tahoma', 13), fg_color="#005369", hover_color="#008097")
    run_button.place(x=170, y=200)

    app.mainloop()