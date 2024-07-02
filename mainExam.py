#Cào dữ liệu khám
#using
import csv
import threading
from tkinter import Menu
import os
import time
import tkinter as tk
from tkinter import messagebox
from datetime import datetime, timedelta
import customtkinter
from tkinter import filedialog
from seleniumwire import webdriver
from selenium.webdriver.common.by import By
from PIL import ImageTk, Image
from tkcalendar import Calendar
from babel.numbers import *
import json
import pandas as pd
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from tkinter import ttk
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

#global
customtkinter.set_appearance_mode("Light")
customtkinter.set_default_color_theme("green")
dataGender = []
dataStatus = []
dataShift = []
pathExcel = ""

#class
class ExcelModel:
    dateKham = ""
    listCode = []

class PatientModel:
    stt = 0
    code = ""
    name = ""
    datebirth = ""
    gender = ""
    age = 0
    province = ""
    district = ""
    ward = ""
    date = ""
    shift = ""
    primaryICD10Code = ""
    primaryICD10 = ""
    secondICD10 = ""
    diagnosis = ""
    status = ""

    def to_dict(self):
        return {
            "STT": self.stt,
            "Mã bệnh nhân": self.code,
            "Tên bệnh nhân": self.name,
            "Ngày sinh": self.datebirth,
            "Giới tính": self.gender,
            "Tuổi": self.age,
            "Tỉnh Thành": self.province,
            "Quận Huyện": self.district,
            "Xã Phường": self.ward,
            "Ngày khám": self.date,
            "Ca khám": self.shift,
            "Mã ICD chính": self.primaryICD10Code,
            "Tên ICD chính": self.primaryICD10,
            "Mã ICD phụ": self.secondICD10,
            "Chuẩn đoán 2": self.diagnosis,
            "Tình trạng": self.status,
        }

    def printModel(self):
        print("-------------------------------------------------")
        print(f"STT: {self.stt}")
        print(f"Mã bệnh nhân: {self.code}")
        print(f"Tên bệnh nhân: {self.name}")
        print(f"Ngày sinh: {self.datebirth}")
        print(f"Tuổi: {self.age}")
        print(f"Giới tính: {self.gender}")
        print(f"Phường/Xã: {self.ward}")
        print(f"Quận/Huyện: {self.district}")
        print(f"Tỉnh/Thành Phố: {self.province}")
        print(f"Ngày khám: {self.date}")
        print(f"Ca Khám: {self.shift}")
        print(f"ICD1: {self.primaryICD10Code}")
        print(f"Tên ICD1: {self.primaryICD10}")
        print(f"ICD 2 (phụ): {self.secondICD10}")
        print(f"Chuẩn đoán 2: {self.diagnosis}")
        print(f"Tình trạng: {self.status}")
        print("-------------------------------------------------")

    def ExportModel(self):
        headers = [
            ("STT:", self.stt),
            ("Mã bệnh nhân:", self.code),
            ("Tên bệnh nhân:", self.name),
            ("Ngày sinh:", self.datebirth),
            ("Tuổi:", self.age),
            ("Giới tính:", self.gender),
            ("Phường/Xã:", self.ward),
            ("Quận/Huyện:", self.district),
            ("Tỉnh/Thành Phố:", self.province),
            ("Ngày khám:", self.date),
            ("Ca Khám:", self.shift),
            ("ICD1:", self.primaryICD10Code),
            ("Tên ICD1:", self.primaryICD10),
            ("ICD2 (phụ):", self.secondICD10),
            ("Chuẩn đoán 2:", self.diagnosis),
            ("Tình trạng:", self.status)
        ]

        # Tính toán độ rộng cột
        max_key_length = max(len(header[0]) for header in headers) + 2
        max_value_length = max(len(str(header[1])) for header in headers) + 2

        s = ""
        s += "+" + "-" * max_key_length + "+" + "-" * max_value_length + "+\n"

        for header, value in headers:
            key_column = f" {header:<{max_key_length-2}} "
            value_column = f" {value:<{max_value_length-2}} "
            s += f"|{key_column}|{value_column}|\n"

        s += "+" + "-" * max_key_length + "+" + "-" * max_value_length + "+\n"
        return s
        
class GenderPatientModel:
    # thuộc tính
    gender_code = ""
    gender_name = ""
    # phương thức
    def setName(self, gender_name):
        self.gender_name = gender_name
    
    def getName(self):
        return self.gender_name
    
    def setCode(self, gender_code):
        self.gender_code = gender_code
    
    def getCode(self):
        return self.gender_code

class StatusPatientModel:
    # thuộc tính
    status_code = ""
    status_name = ""
    # phương thức
    def setName(self, status_name):
        self.status_name = status_name
    
    def getName(self):
        return self.status_name
    
    def setCode(self, status_code):
        self.status_code = status_code
    
    def getCode(self):
        return self.status_code
    
class Shift:
    shift_id = 0
    shift_name = ""

# class DatePicker(tk.Toplevel):
#     def __init__(self, parent, date_entry, next_widget):
#         super().__init__(parent)
#         self.title("Date Picker")
#         self.overrideredirect(True)
#         self.grab_set()
#         self.date_entry = date_entry
#         self.next_widget = next_widget

#         self.cal = Calendar(self, selectmode="day", year=current_date.year, month=current_date.month,
#                             day=current_date.day)
#         self.cal.pack()

#         confirm_button = customtkinter.CTkButton(self, text="Confirm", command=self.update_date)
#         confirm_button.pack()

#     def update_date(self):
#         selected_date_str = self.cal.get_date()
#         selected_date = datetime.datetime.strptime(selected_date_str, "%m/%d/%y")  # Convert to datetime object
#         self.date_entry.delete(0, tk.END)
#         self.date_entry.insert(0, selected_date.strftime("%d/%m/%Y"))
#         self.destroy()
#         run_button.focus_set()

#funtion
def center_window(window, width=600, height=440):
    # Lấy kích thước màn hình
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    # Tính toán tọa độ x và y để cửa sổ ở giữa màn hình
    x = (screen_width / 2) - (width / 2)
    y = (screen_height / 2) - (height / 2)
    window.geometry(f'{width}x{height}+{int(x)}+{int(y)}')

def open_secondary_interface():
    app.withdraw()  # Ẩn cửa sổ chính
    import mainReportExam as secondary  # Import file secondary.py và mở giao diện mới
    secondary.run_secondary_interface(app)

def open_presciption_interface():
    app.withdraw()  # Ẩn cửa sổ chính
    import mainPrescription as scription  # Import file secondary.py và mở giao diện mới
    scription.run_secondary_interface(app)

def clear_requests(driver):
    driver.requests.clear()

def is_within_period(date_string):
    dt = datetime.datetime.strptime(date_string, "%Y-%m-%dT%H:%M:%S.%fZ")
    start_period = dt.replace(hour=21, minute=0, second=0, microsecond=0)
    end_period = dt.replace(hour=23, minute=59, second=59, microsecond=999999)
    if start_period <= dt <= end_period:
        correct_date = (dt + timedelta(days=1)).strftime("%Y-%m-%d")
    else:
        correct_date = dt.strftime("%Y-%m-%d")
    return correct_date

def change_year(date,yy):
    if isinstance(date, str):
        dt = datetime.datetime.strptime(date, "%y-%m-%d")
    elif isinstance(date, datetime.date):
        dt = date
    else:
        raise ValueError("Unsupported date format")
    
    try:
        new_year = int(yy)
    except ValueError:
        raise ValueError("Invalid year format")
    
    new_date = dt.replace(year=new_year)
    return new_date

def pick_year(val):
    global yearCheck
    yearCheck = val

def loadDateEntry():
    current_year = datetime.datetime.now().year
    years = list(range(current_year, 0, -1))
    return years

def find_shift(id):
    for s in dataShift:
        if s.shift_id == id:
            return s.shift_name
    return None

def find_gender(id):
    for s in dataGender:
        if s.getCode() == id:
            return s.getName()
    return None

def find_status(id):
    for s in dataStatus:
        if s.getCode() == id:
            return s.getName()
    return None

def filer_status(strStatus):
    if strStatus == "Nhập viện":
        return find_status("nhapvien")
    
    if strStatus == "Chuyển viện":
        return find_status("chuyenvien")
    
    if strStatus == "Khác":
        return find_status("khac")
    
    return find_status("toave")

# def pick_date(event):
#     next_widget = event.widget.tk_focusNext()
#     date_picker = DatePicker(app, date_entry, next_widget)
#     x, y, _, _ = event.widget.bbox("insert")
#     x += event.widget.winfo_rootx() - 7
#     y += event.widget.winfo_rooty() + 30
#     date_picker.geometry("+%d+%d" % (x, y))

def validate_input():
    if pathExcel == "":
        return False
    return True

def select_Excel_file():
    global pathExcel
    # Hiển thị hộp thoại chọn file PDF
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if file_path:
        # Lấy đường dẫn file
        pathExcel = file_path
        # Lấy tên file từ đường dẫn
        file_name = os.path.basename(file_path)
        # Cập nhật text của nút với tên file
        drop.configure(text=file_name)
        # Cập nhật fg_color của nút
        drop.configure(fg_color="#5DBD60")
        # Tắt enable nút
        drop.configure(state="disabled")

def cancel_Excel_file():
    global pathExcel
    if pathExcel == '':
        messagebox.showerror(title="Thất bại!", message="Chưa chọn file nào!")
    else:
        pathExcel = ''
        drop.configure(text="Ấn để chọn file Excel!")
        # Cập nhật fg_color của nút
        drop.configure(fg_color="#FF9292")
        # Bật enable nút
        drop.configure(state="enable")

def loadingData():
    gd1 = GenderPatientModel()
    gd2 = GenderPatientModel()
    gd1.setCode("male")
    gd2.setCode("female")
    gd1.setName("NAM")
    gd2.setName("NỮ")
    dataGender.append(gd1)
    dataGender.append(gd2)

    st1 = StatusPatientModel()
    st2 = StatusPatientModel()
    st3 = StatusPatientModel()
    st4 = StatusPatientModel()
    st1.setCode("toave")
    st2.setCode("nhapvien")
    st1.setName("TOA VỀ")
    st2.setName("NHẬP VIỆN")
    st3.setCode("chuyenvien")
    st4.setCode("khac")
    st3.setName("CHUYỂN VIỆN")
    st4.setName("KHÁC")
    dataStatus.append(st1)
    dataStatus.append(st2)
    dataStatus.append(st3)
    dataStatus.append(st4)

    s1 = Shift()
    s1.shift_id = 0
    s1.shift_name = "Ca 0"
    s2 = Shift()
    s2.shift_id = 1
    s2.shift_name = "Ca 1"
    s3 = Shift()
    s3.shift_id = 2
    s3.shift_name = "Ca 2"
    s4 = Shift()
    s4.shift_id = 3
    s4.shift_name = "Ca 3"
    s5 = Shift()
    s5.shift_id = 4
    s5.shift_name = "Ca 4"
    s6 = Shift()
    s6.shift_id = 5
    s6.shift_name = "Ca 5"
    s7 = Shift()
    s7.shift_id = 6
    s7.shift_name = "Ca 6"
    s8 = Shift()
    s8.shift_id = 7
    s8.shift_name = "Ca HC"
    s9 = Shift()
    s9.shift_id = 8
    s9.shift_name = "Ca 0 - Sau KK"
    dataShift.append(s1)
    dataShift.append(s2)
    dataShift.append(s3)
    dataShift.append(s4)
    dataShift.append(s5)
    dataShift.append(s6)
    dataShift.append(s7)
    dataShift.append(s8)
    dataShift.append(s9)

def convert_date_format(date_str):
    try:
        date_obj = datetime.datetime.strptime(date_str, "%d/%m/%Y")
        return date_obj.strftime("%Y-%m-%d")
    except ValueError:
        print(f"Invalid date format: {date_str}")
        return date_str

def start_script_thread(listmodels, directory):
    global script_thread  # Khai báo biến toàn cục
    terminal_window, terminal_text = open_terminal_window()
    app.withdraw()
    script_thread = threading.Thread(target=run_script, args=(listmodels, directory, terminal_text))
    script_thread.start()

def open_terminal_window():
    terminal_window = tk.Toplevel(app)
    terminal_window.title("Màn hình chạy dữ liệu")
    terminal_window.attributes("-fullscreen", True)
    icon_path = os.path.abspath("resource\crawlLogo.ico")
    terminal_window.iconbitmap(icon_path)
    terminal_text = tk.Text(terminal_window, bg="black", fg="green", insertbackground="green")
    terminal_text.pack(expand=True, fill='both')

    def on_closing():
        terminal_window.destroy()
        app.deiconify()

    terminal_window.protocol("WM_DELETE_WINDOW", on_closing)
    return terminal_window, terminal_text


def run_script(listmodels,directory,terminal_text):
    global pathExcel, yearCheck

    # Hàm để hiển thị thông tin lên terminal
    def log_terminal(message):
        terminal_text.insert(tk.END,message + "\n")
        terminal_text.see(tk.END)  # Scroll to the end
        terminal_text.update_idletasks()
    #duyệt danh sách model
    for model in listmodels:
        dateKham = model['date']
        dateKham = dateKham.strftime("%Y-%m-%d")
        patient_codes_list = model['patient_codes']
        prescription_list = model['prescriptions_id']
        patient_result_list = []
        #csv file
        csv_filename = os.path.join(directory, f"Patient_Data_{dateKham}.csv")
        # CSV file header
        csv_header = ["STT", "Mã bệnh nhân", "Tên bệnh nhân", "Ngày sinh", "Giới tính", "Tuổi", "Tỉnh Thành", "Quận Huyện", "Xã Phường", "Ngày khám", "Ca khám", "Mã ICD chính", "Tên ICD chính", "Mã ICD phụ", "Chuẩn đoán 2", "Tình trạng"]
         # Kiểm tra nếu file CSV đã tồn tại và đếm số dòng
        start_index = 0
        if os.path.exists(csv_filename):
            with open(csv_filename, mode='r', encoding='utf-8-sig') as file:
                reader = csv.reader(file)
                start_index = sum(1 for row in reader) - 1  # Trừ đi header
        demstt = start_index
        breakSTT = 0
        for pt,pl in zip(patient_codes_list[start_index:],prescription_list[start_index:]):
            success = False
            errorDrug = 0
            errorLogWeb = 0 
            while not success:
                try:
                    if breakSTT == 0 and errorDrug == 0:
                        #chrome settings
                        chrome_options = webdriver.ChromeOptions()
                        chrome_options.add_experimental_option("prefs", {
                            "download.default_directory": directory,
                            "profile.default_content_setting_values.autom   atic_downloads": 1,
                            "download.prompt_for_download": False,
                            "profile.default_content_settings.popups": 0,
                            "safebrowsing.enabled": "false",
                            "safebrowsing.disable_download_protection": True
                        })
                        ascii_nd2 = """
........................Đang kết nối lại website..............................
                        """
                        log_terminal(ascii_nd2)
                        prefs = {"credentials_enable_service": False,
                            "profile.password_manager_enabled": False}
                        chrome_options.add_experimental_option("prefs", prefs)
                        chrome_options.add_argument("--headless")  # Chạy trình duyệt trong chế độ headless
                        chrome_options.add_argument("--disable-gpu")  # Tăng tốc độ trên các hệ điều hành không có GPU
                        chrome_options.add_argument("--window-size=1920x1080")  # Thiết lập kích thước cửa sổ mặc định
                        driver = webdriver.Chrome(options=chrome_options)
                        url = "http://192.168.0.77/dist/#!/login"
                        try:
                            driver.get(url)
                        except Exception as e:
                            errorLogWeb += 1
                            log_terminal(f"Lỗi khi truy cập URL {url}: {e}")
                            if errorLogWeb >= 10:
                                log_terminal("Lỗi server, dừng chương trình.")
                                return
                            time.sleep(3)  # Chờ một chút trước khi thử lại
                            continue

                        errorLogWeb = 0  # Reset biến đếm lỗi khi truy cập URL thành công
                        time.sleep(3)
                        log_terminal(".........................Duyệt WEBSITE thành công.................................")
                        # login
                        form = driver.find_element(By.NAME, "LoginForm")
                        username = driver.find_element(By.NAME, 'username')
                        password = driver.find_element(By.NAME, 'password')
                        username.send_keys('quyen.ngoq')
                        password.send_keys('74777477')
                        # ấn nút login
                        form.submit()
                        time.sleep(2)
                        # Chọn
                        liElement = driver.find_element(By.XPATH,
                                                        "//li[contains(@class, 'nav-item pl-2 pr-3 d-lg-none d-xl-block') and contains(@permission, "
                                                        "'patient_info_payment')]")
                        liElement.click()
                        time.sleep(1)
                    cod = str(pt)
                    patient_result = PatientModel()
                    patientcode = driver.find_element(By.NAME, 'patient_code_qr')
                    patientcode.clear()
                    patientcode.send_keys(cod)
                    time.sleep(0.1)
                    btnFind = driver.find_element(By.XPATH, "//button[contains(text(), 'Tìm')]")
                    btnFind.click()
                    time.sleep(0.1)
                    for request in reversed(driver.requests):
                        if request.response:
                            if 'load_patientById' in request.url:
                                # format cho dữ liệu về dạng json
                                try:
                                    patient_data = json.loads(request.response.body)
                                    # đọc data
                                    patient_info = patient_data.get('data', {}).get('patient', {})
                                    if patient_info.get('patient_code') == cod:
                                        patient_result.code = patient_info.get('patient_code')
                                        patient_result.name = patient_info.get('name')
                                        patient_result.datebirth = patient_info.get('date_of_birth')
                                        patient_result.gender = find_gender(patient_info.get('gender'))
                                        patient_result.age = patient_info.get('age')
                                        patient_result.province = patient_info.get('province')
                                        patient_result.district = patient_info.get('district')
                                        patient_result.ward = patient_info.get('ward_name')
                                        if patient_result.district == None:
                                            patient_result.district = ''
                                        if patient_result.ward == None:
                                            patient_result.ward = ''
                                        break
                                except:
                                    continue
                    btnThuoc = driver.find_element(By.XPATH, "//a[contains(text(), 'Thuốc Ngoại Trú ')]")
                    btnThuoc.click()
                    time.sleep(0.1)
                    for request in reversed(driver.requests):
                        foundDrug = False
                        if request.response:
                            if 'getListPrescriptionOutpatient' in request.url:
                                # format cho dữ liệu về dạng json
                                try:
                                    drug_data = json.loads(request.response.body)
                                    # đọc data
                                    drug_info = drug_data.get('data', {})
                                    for di in drug_info:
                                        if  di.get('check_in_out_hospital_record_id') == 0:
                                            if di.get('prescription_id') == int(pl):
                                                patient_result.date = dateKham
                                                patient_result.shift = find_shift(di.get('shift_id'))
                                                patient_result.primaryICD10Code = di.get('primary_icd10_code')
                                                patient_result.primaryICD10 = di.get('primary_icd10')
                                                patient_result.secondICD10 = di.get('second_icd10_code')
                                                patient_result.diagnosis = di.get('diagnosis')
                                                patient_result.status = filer_status(di.get('note'))
                                                foundDrug = True
                                                break
                                except:
                                    continue     
                        if foundDrug == True:
                            break
                    if patient_result.code == "" or patient_result.primaryICD10 == "":
                        errorDrug += 1
                        log_terminal(f"Dữ liệu bệnh nhân trống, mã bệnh nhân {cod}, đang cố thử lại...")
                        if errorDrug >= 10:
                            log_terminal(f"Quá nhiều lỗi dữ liệu trống, bỏ qua mã bệnh nhân {cod}")
                            success = True  # Bỏ qua mã bệnh nhân hiện tại
                        continue  # Thử lại mã bệnh nhân hiện tại
                    else:
                        errorDrug = 0  # Reset biến đếm lỗi khi dữ liệu hợp lệ
                        success = True
                        demstt = demstt + 1
                        patient_result.stt = demstt
                        patient_result_list.append(patient_result)
                        log_terminal(patient_result.ExportModel())
                        if demstt % 100 == 0:
                            # Ghi dữ liệu tạm thời vào file CSV
                            with open(csv_filename, mode='a', newline='', encoding='utf-8') as file:
                                writer = csv.DictWriter(file, fieldnames=csv_header)
                                if file.tell() == 0:  # Nếu file rỗng, ghi header
                                    writer.writeheader()
                                writer.writerows([pt.to_dict() for pt in patient_result_list])
                            breakSTT = 0
                            driver.quit()
                            patient_result_list.clear()
                            log_terminal(".........................Ghi tạm vào file CSV........................................")
                        else:
                            breakSTT = breakSTT + 1
                except:
                    success = False
        # Ghi dữ liệu còn lại vào file CSV nếu có
        if patient_result_list:
            with open(csv_filename, mode='a', newline='', encoding='utf-8') as file:
                writer = csv.DictWriter(file, fieldnames=csv_header)
                if file.tell() == 0:
                    writer.writeheader()
                writer.writerows([pt.to_dict() for pt in patient_result_list])
        # Thoát trình duyệt 
        driver.quit()
        # Chuyển dữ liệu từ file CSV sang file Excel
        try:
            df = pd.read_csv(csv_filename, encoding='utf-8-sig',dtype={'Mã bệnh nhân': str})
            output_filename = os.path.join(directory, f"Patient_Data_{dateKham}.xlsx")
            with pd.ExcelWriter(output_filename, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name='Sheet1')

            # Định dạng file Excel
            wb = load_workbook(output_filename)
            ws = wb.active

            # Định dạng header
            header_font = Font(bold=True)
            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Tự động điều chỉnh độ rộng cột
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            
            # Định dạng cột Mã bệnh nhân là Text
            for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
                for cell in row:
                    cell.number_format = '@'
            
            wb.save(output_filename)
            os.remove(csv_filename)  # Xóa file CSV sau khi chuyển sang Excel
            log_terminal(f"Quá trình thu thập dữ liệu khám ngày {dateKham} thành công!")
        except Exception as e:
            messagebox.showerror(title="Lỗi", message=f"Lỗi quá trình tạo file excel không thành công! {e}")  
    
    # Đóng cửa sổ terminal và hiển thị lại cửa sổ chính
    app.deiconify()

def handle_file_excel():
    #khai báo biến

    #xét lỗi và cho chọn file excel
    valid = validate_input()
    if valid == False:
        messagebox.showerror(title="Lỗi dữ liệu", message="Chưa chọn file excel dữ liệu")
        return
    
    directory = filedialog.askdirectory(title="Chọn nơi lưu trữ!")
    if directory:
        directory = directory.replace('/', '\\')
    else:
        return
    
    existing_excel_files = [f for f in os.listdir(directory) if f.endswith('.xlsx') or f.endswith('.xls')]
    num_existing_files = len(existing_excel_files)

    listmodel = []
    # Đọc file excel
    if pathExcel.endswith('.xlsx'):
        excel_file = pd.ExcelFile(pathExcel, engine='openpyxl')
    elif pathExcel.endswith('.xls'):
        excel_file = pd.ExcelFile(pathExcel, engine='xlrd')
    else:
        raise ValueError("Unsupported file format")

    # Duyệt qua các sheet của file excel đầu vào và bỏ qua số sheet tương ứng với số file Excel đã có trong thư mục
    for idx, sheet_name in enumerate(excel_file.sheet_names):
        if idx < num_existing_files:
            continue  # Bỏ qua sheet này
        try:
            # Cố gắng chuyển đổi tên sheet thành định dạng ngày
            sheet_name_trip = sheet_name.strip()
            date = datetime.datetime.strptime(sheet_name_trip, "%d-%m-%Y").date()
        except ValueError:
            # Nếu không thể chuyển đổi, in ra thông báo lỗi và bỏ qua sheet này
            print(f"Tên sheet '{sheet_name}' không đúng định dạng ngày. Sheet này sẽ bị bỏ qua.")
            continue

        df = excel_file.parse(sheet_name)
        patient_codes = df.iloc[2:, 1]  # Đọc cột mã bệnh nhân từ dòng thứ 3
        precriptions_id = df.iloc[2:, 3]  # Đọc cột trạng thái khám từ dòng thứ 3

        # Chỉ lấy những dòng là mã bệnh nhân
        valid_patient_codes = []
        valid_patient_precsriptions =[]
        for a, b in zip(patient_codes, precriptions_id):
            code_str = str(a)  # Convert to string
            precripstr = str(b)  # Convert to string
            if code_str[0].isdigit():
                if precripstr != "nan":
                    valid_patient_codes.append(code_str)
                    valid_patient_precsriptions.append(precripstr)                   
        
        # Nếu không có danh sách mã bệnh nhân hợp lệ thì next
        if not valid_patient_codes:
            continue

        # Tạo model
        model = {
            'date': date,
            'patient_codes': valid_patient_codes,
            'prescriptions_id': valid_patient_precsriptions
        }
        listmodel.append(model)

    start_script_thread(listmodel, directory)
#app
app = customtkinter.CTk()
app.title("Lấy dữ liệu khám chữa bệnh")
center_window(app)
icon_path = os.path.abspath("resource\crawlLogo.ico")
app.iconbitmap(icon_path)

# Tạo menu bar
menu_bar = Menu(app)
app.config(menu=menu_bar)

# Thêm mục vào menu bar
report_menu = Menu(menu_bar, tearoff=0)
menu_bar.add_cascade(label="Xuất báo cáo", menu=report_menu)
report_menu.add_command(label="Xuất báo cáo toa khám bệnh", command=open_presciption_interface)
report_menu.add_command(label="Xuất báo cáo đợt khám bệnh", command=open_secondary_interface)

detail_menu = Menu(menu_bar, tearoff=0)
menu_bar.add_cascade(label="Chi tiết khám", menu=detail_menu)
detail_menu.add_command(label="Chi tiết toa thuốc", command=open_presciption_interface)
detail_menu.add_command(label="Chi tiết chỉ định", command=open_secondary_interface)

loadingData()

imgBG = ImageTk.PhotoImage(Image.open("resource\BG.png"))
l1 = customtkinter.CTkLabel(master=app, image=imgBG)
l1.pack()

frame = customtkinter.CTkFrame(master=l1, width=320, height=300, corner_radius=15)
frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

l2 = customtkinter.CTkLabel(master=frame, text="Select configuration", font=('Century Gothic', 20))
l2.place(x=70, y=45)

drop = customtkinter.CTkButton(master=frame, command=select_Excel_file, width=300, height=50, corner_radius=15, fg_color="#FF9292", hover_color="#FFB6B6", text_color="#000000", text="Ấn để chọn file Excel!", font=('Tahoma', 13))
drop.place(x=10, y=100)

cancel_button = customtkinter.CTkButton(master=frame, text="Hủy bỏ", command=cancel_Excel_file, font=('Tahoma', 13), fg_color="#B3271C", hover_color="#FF3D4D")
cancel_button.place(x=10, y=200)

run_button = customtkinter.CTkButton(master=frame, text="Thực thi", command=handle_file_excel, font=('Tahoma', 13), fg_color="#005369", hover_color="#008097")
run_button.place(x=170, y=200)

def on_closing():
    app.quit()
app.protocol("WM_DELETE_WINDOW", on_closing)
app.mainloop()