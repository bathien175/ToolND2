#cào hóa đơn và dịch vụ đã sử dụng
#using
import csv
import os
import time
import tkinter as tk
from tkinter import messagebox
from datetime import datetime, timedelta
import customtkinter
from tkinter import filedialog
from seleniumwire import webdriver
from selenium.webdriver.common.by import By
from PIL import ImageTk, Image
from tkcalendar import Calendar
from babel.numbers import *
import json
import pandas as pd
from tkinter import ttk
from openpyxl.cell import MergedCell
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
import threading
#global
customtkinter.set_appearance_mode("Light")
customtkinter.set_default_color_theme("green")
pathExcel = ""
drop = Any
radio_var = Any
#class
class InvoiceDetailModel:
    stt = 0
    invoice_code = ""
    pay_payment_item_id = 0
    pay_receipt_id = 0
    patient_id = 0
    cashier_id = 0
    cashier_name = ""
    ngay_thu = ""
    nguoi_thu = ""
    doctor_name = ""
    created_time = ""
    e_invoice_id = 0
    e_invoice_status = 0
    enum_invoice_type = 0
    enum_item_type = 0
    enum_payment_type = 0
    hd_date = ""
    hd_info = ""
    da_thu = ""
    deleted = 0
    returned = 0
    refund_date = ""
    refund_amount = 0
    refund_invoice_code = ""
    refund_pay_receipt_id = 0
    new_invoice_code = ""
    service_id = 0
    code = ""
    service_name = ""
    insurance_name = ""
    item_type = ""
    unit = ""
    unit_price = 0
    quantity = 0
    total = 0
    total_dichvu = 0
    total_nutrition = 0
    total_price = 0
    patient_price = 0
    total_returned_amount = 0
    tam_ung = 0
    phai_thu = 0
    phai_tra = 0
    so_hd = ""
    so_luu_tru = ""
    shift = 0
    zone = ""
    note = ""
    IDMapping = 0
    def to_dict(self):
        return {
            "STT": self.stt,
            "Mã hóa đơn": self.invoice_code,
            "pay_payment_item_id": self.pay_payment_item_id,
            "pay_receipt_id": self.pay_receipt_id,
            "ID bệnh nhân": self.patient_id,
            "ID thu ngân": self.cashier_id,
            "Tên thu ngân": self.cashier_name,
            "Ngày thu": self.ngay_thu,
            "Người thu": self.nguoi_thu,
            "Tên bác sĩ": self.doctor_name,
            "Thời gian tạo": self.created_time,
            "Mã hóa đơn điện tử": self.e_invoice_id,
            "Trạng thái hóa đơn điện tử": self.e_invoice_status,
            "Loại hóa đơn": self.enum_invoice_type,
            "Phân loại": self.enum_item_type,
            "Loại thanh toán": self.enum_payment_type,
            "Ngày hóa đơn": self.hd_date,
            "Thông tin hóa đơn": self.hd_info,
            "Đã thu": self.da_thu,
            "Đã xóa": self.deleted,
            "Hoàn trả": self.returned,
            "refund_date": self.refund_date,
            "refund_amount": self.refund_amount,
            "refund_invoice_code": self.refund_invoice_code,
            "refund_pay_receipt_id": self.refund_pay_receipt_id,
            "new_invoice_code": self.new_invoice_code,
            "ID dịch vụ": self.service_id,
            "Mã dịch vụ": self.code,
            "Tên dịch vụ": self.service_name,
            "Tên bảo hiểm": self.insurance_name,
            "Loại dịch vụ": self.item_type,
            "Đơn vị tính": self.unit,
            "Đơn giá": self.unit_price,
            "Số lượng": self.quantity,
            "Total": self.total,
            "total_dichvu": self.total_dichvu,
            "total_nutrition": self.total_nutrition,
            "total_price": self.total_price,
            "patient_price": self.patient_price,
            "total_returned_amount": self.total_returned_amount,
            "tam_ung": self.tam_ung,
            "phai_thu": self.phai_thu,
            "phai_tra": self.phai_tra,
            "so_hd": self.so_hd,
            "so_luu_tru": self.so_luu_tru,
            "shift": self.shift,
            "Khu vực": self.zone,
            "Ghi chú": self.note,
            "IDMapping": self.IDMapping
        }
    def ExportModel(self):
        headers = [
            ("STT", self.stt),
            ("Mã hóa đơn", self.invoice_code),
            ("pay_payment_item_id", self.pay_payment_item_id),
            ("pay_receipt_id", self.pay_receipt_id),
            ("ID bệnh nhân", self.patient_id),
            ("ID thu ngân", self.cashier_id),
            ("Tên thu ngân", self.cashier_name),
            ("Ngày thu", self.ngay_thu),
            ("Người thu", self.nguoi_thu),
            ("Tên bác sĩ", self.doctor_name),
            ("Thời gian tạo", self.created_time),
            ("Mã hóa đơn điện tử", self.e_invoice_id),
            ("Trạng thái hóa đơn điện tử", self.e_invoice_status),
            ("Loại hóa đơn", self.enum_invoice_type),
            ("Phân loại", self.enum_item_type),
            ("Loại thanh toán", self.enum_payment_type),
            ("Ngày hóa đơn", self.hd_date),
            ("Thông tin hóa đơn", self.hd_info),
            ("Đã thu", self.da_thu),
            ("Đã xóa", self.deleted),
            ("Hoàn trả", self.returned),
            ("refund_date", self.refund_date),
            ("refund_amount", self.refund_amount),
            ("refund_invoice_code", self.refund_invoice_code),
            ("refund_pay_receipt_id", self.refund_pay_receipt_id),
            ("new_invoice_code", self.new_invoice_code),
            ("ID dịch vụ", self.service_id),
            ("Mã dịch vụ", self.code),
            ("Tên dịch vụ", self.service_name),
            ("Tên bảo hiểm", self.insurance_name),
            ("Loại dịch vụ", self.item_type),
            ("Đơn vị tính", self.unit),
            ("Đơn giá", self.unit_price),
            ("Số lượng", self.quantity),
            ("Total", self.total),
            ("total_dichvu", self.total_dichvu),
            ("total_nutrition", self.total_nutrition),
            ("total_price", self.total_price),
            ("patient_price", self.patient_price),
            ("total_returned_amount", self.total_returned_amount),
            ("tam_ung", self.tam_ung),
            ("phai_thu", self.phai_thu),
            ("phai_tra", self.phai_tra),
            ("so_hd", self.so_hd),
            ("so_luu_tru", self.so_luu_tru),
            ("shift", self.shift),
            ("Khu vực", self.zone),
            ("Ghi chú", self.note),
            ("IDMapping", self.IDMapping),
        ]

        # Tính toán độ rộng cột
        max_key_length = max(len(header[0]) for header in headers) + 2
        max_value_length = max(len(str(header[1])) for header in headers) + 2

        s = ""
        s += "+" + "-" * max_key_length + "+" + "-" * max_value_length + "+\n"

        for header, value in headers:
            key_column = f" {header:<{max_key_length-2}} "
            value_column = f" {value:<{max_value_length-2}} "
            s += f"|{key_column}|{value_column}|\n"

        s += "+" + "-" * max_key_length + "+" + "-" * max_value_length + "+\n"
        return s

class InvoiceModel:
    stt = 0
    patient_code = ""
    invoice_code = ""
    bhyt = 0
    created_time = ""
    cashier_name = ""
    enum_examination_type = 0
    enum_invoice_type = 0
    enum_patient_type = 0
    enum_payment_type = 0
    pay_receipt_id = 0
    ticket_id = 0
    so_hd = ""
    so_luu_tru = ""
    total_price = 0
    patient_price = 0
    total_returned_amount = 0
    returned = 0
    deleted = 0
    note = ""
    IDMapping = 0
    def to_dict(self):
        return {
            "STT": self.stt,
            "Mã bệnh nhân": self.patient_code,
            "Mã hóa đơn": self.invoice_code,
            "Bảo hiểm y tế": self.bhyt,
            "Thời gian tạo": self.created_time,
            "Tên thu ngân": self.cashier_name,
            "Loại dịch vụ khám": self.enum_examination_type,
            "Loại hóa đơn": self.enum_invoice_type,
            "Loại bệnh nhân": self.enum_patient_type,
            "Hình thức thanh toán": self.enum_payment_type,
            "pay_receipt_id": self.pay_receipt_id,
            "Mã đợt khám": self.ticket_id,
            "Số hóa đơn": self.so_hd,
            "Số lưu trữ": self.so_luu_tru,
            "Tổng tiền": self.total_price,
            "Bệnh nhân trả": self.patient_price,
            "Tiền hoàn lại": self.total_returned_amount,
            "Hoàn trả hóa đơn": self.returned,
            "Xóa": self.deleted,
            "Ghi chú": self.note,
            "IDMapping": self.IDMapping
        }
    def ExportModel(self):
        headers = [
            ("STT", self.stt),
            ("Mã bệnh nhân", self.patient_code),
            ("Mã hóa đơn", self.invoice_code),
            ("Bảo hiểm y tế", self.bhyt),
            ("Thời gian tạo", self.created_time),
            ("Tên thu ngân", self.cashier_name),
            ("Loại dịch vụ khám", self.enum_examination_type),
            ("Loại hóa đơn", self.enum_invoice_type),
            ("Loại bệnh nhân", self.enum_patient_type),
            ("Loại thanh toán", self.enum_payment_type),
            ("pay_receipt_id", self.pay_receipt_id),
            ("Mã đợt khám", self.ticket_id),
            ("Số hóa đơn", self.so_hd),
            ("Số lưu trữ", self.so_luu_tru),
            ("Tổng tiền", self.total_price),
            ("Bệnh nhân trả", self.patient_price),
            ("Tiền hoàn lại", self.total_returned_amount),
            ("Hoàn trả hóa đơn", self.returned),
            ("Xóa", self.deleted),
            ("Ghi chú", self.note),
            ("IDMapping", self.IDMapping)
        ]

        # Tính toán độ rộng cột
        max_key_length = max(len(header[0]) for header in headers) + 2
        max_value_length = max(len(str(header[1])) for header in headers) + 2

        s = ""
        s += "+" + "-" * max_key_length + "+" + "-" * max_value_length + "+\n"

        for header, value in headers:
            key_column = f" {header:<{max_key_length-2}} "
            value_column = f" {value:<{max_value_length-2}} "
            s += f"|{key_column}|{value_column}|\n"

        s += "+" + "-" * max_key_length + "+" + "-" * max_value_length + "+\n"
        return s

#function
def get_last_value_of_column(csv_filename, column_index):
    last_value = None
    with open(csv_filename, mode='r', encoding='utf-8-sig') as file:
        reader = csv.reader(file)
        for row in reader:
            if len(row) > column_index:
                last_value = row[column_index]
    return last_value

def get_unit(key):
    df = pd.read_csv("resource\enum_unit_usage.csv")
    # Tìm giá trị dựa trên khóa
    value = df.loc[df['key'] == key, 'value']
    if not value.empty:
        return value.iloc[0]
    else:
        return None

def check_valid_used(key):
    if key == "" or key == "0":
        return "0"
    else:
        return key

def start_script_thread_Invoice(listmodels, directory):
    global script_thread, radio_var  # Khai báo biến toàn cục
    terminal_window, terminal_text = open_terminal_window()
    app.withdraw()
    if radio_var.get() == 0:
        script_thread = threading.Thread(target=run_script_Invoice, args=(listmodels, directory, terminal_text))
    else:
        script_thread = threading.Thread(target=run_script_InvoiceDetail, args=(listmodels, directory, terminal_text))
    script_thread.start()

def open_terminal_window():
    terminal_window = tk.Toplevel(app)
    terminal_window.title("Màn hình chạy dữ liệu")
    terminal_window.attributes("-fullscreen", True)
    icon_path = os.path.abspath("resource\crawlLogo.ico")
    terminal_window.iconbitmap(icon_path)
    terminal_text = tk.Text(terminal_window, bg="black", fg="green", insertbackground="green")
    terminal_text.pack(expand=True, fill='both')

    def on_closing():
        terminal_window.destroy()
        app.deiconify()

    terminal_window.protocol("WM_DELETE_WINDOW", on_closing)
    return terminal_window, terminal_text

def handle_file_excel():
    global radio_var
    if radio_var.get() == 0:
        handle_file_excel_Invoice()
    else:

        handle_file_excel_InvoiceDetail()

def handle_file_excel_Invoice():
    #khai báo biến

    #xét lỗi và cho chọn file excel
    valid = validate_input()
    if valid == False:
        messagebox.showerror(title="Lỗi dữ liệu", message="Chưa chọn file excel dữ liệu")
        return
    
    directory = filedialog.askdirectory(title="Chọn nơi lưu trữ!")
    if directory:
        directory = directory.replace('/', '\\')
    else:
        return
    
    existing_excel_files = [f for f in os.listdir(directory) if f.endswith('.xlsx') or f.endswith('.xls')]
    num_existing_files = len(existing_excel_files)

    listmodel = []
    # Đọc file excel
    if pathExcel.endswith('.xlsx'):
        excel_file = pd.ExcelFile(pathExcel, engine='openpyxl')
    elif pathExcel.endswith('.xls'):
        excel_file = pd.ExcelFile(pathExcel, engine='xlrd')
    else:
        raise ValueError("Unsupported file format")

    # Duyệt qua các sheet của file excel đầu vào và bỏ qua số sheet tương ứng với số file Excel đã có trong thư mục
    for idx, sheet_name in enumerate(excel_file.sheet_names):
        if idx < num_existing_files:
            continue  # Bỏ qua sheet này
        try:
            # Cố gắng chuyển đổi tên sheet thành định dạng ngày
            sheet_name_trip = sheet_name.strip()
            date = datetime.datetime.strptime(sheet_name_trip, "%d-%m-%Y").date()
        except ValueError:
            # Nếu không thể chuyển đổi, in ra thông báo lỗi và bỏ qua sheet này
            print(f"Tên sheet '{sheet_name}' không đúng định dạng ngày. Sheet này sẽ bị bỏ qua.")
            continue

        df = excel_file.parse(sheet_name)
        patient_codes = df.iloc[2:, 1]  # Đọc cột mã bệnh nhân từ dòng thứ 3
        tickets_id = df.iloc[2:, 6]  # Đọc cột mã đợt khám từ dòng thứ 3

        # Chỉ lấy những dòng là mã bệnh nhân
        valid_patient_codes = []
        valid_patient_tickets =[]
        for a, b in zip(patient_codes, tickets_id):
            code_str = str(a)  # Convert to string
            ticketstr = str(b)  # Convert to string
            if code_str[0].isdigit():
                if ticketstr != "nan":
                    valid_patient_codes.append(code_str)
                    valid_patient_tickets.append(ticketstr)                   
        
        # Nếu không có danh sách mã bệnh nhân hợp lệ thì next
        if not valid_patient_codes:
            continue

        # Tạo model
        model = {
            'date': date,
            'patient_codes': valid_patient_codes,
            'tickets_id': valid_patient_tickets
        }
        listmodel.append(model)

    start_script_thread_Invoice(listmodel, directory)

def handle_file_excel_InvoiceDetail():
    #khai báo biến

    #xét lỗi và cho chọn file excel
    valid = validate_input()
    if valid == False:
        return
    
    directory = filedialog.askdirectory(title="Chọn nơi lưu trữ!")
    if directory:
        directory = directory.replace('/', '\\')
    else:
        return
    
    existing_excel_files = [f for f in os.listdir(directory) if f.endswith('.xlsx') or f.endswith('.xls')]
    num_existing_files = len(existing_excel_files)

    listmodel = []
    # Đọc file excel
    if pathExcel.endswith('.xlsx'):
        excel_file = pd.ExcelFile(pathExcel, engine='openpyxl')
    elif pathExcel.endswith('.xls'):
        excel_file = pd.ExcelFile(pathExcel, engine='xlrd')
    else:
        raise ValueError("Unsupported file format")

    # Duyệt qua các sheet của file excel đầu vào và bỏ qua số sheet tương ứng với số file Excel đã có trong thư mục
    for idx, sheet_name in enumerate(excel_file.sheet_names):
        if idx < num_existing_files:
            continue  # Bỏ qua sheet này
        try:
            # Cố gắng chuyển đổi tên sheet thành định dạng ngày
            sheet_name_trip = sheet_name.strip()
            sheet_name_date = datetime.datetime.strptime(sheet_name_trip, "%Y-%m-%d").date()
            dateformat = sheet_name_date.strftime("%d-%m-%Y")
            date = datetime.datetime.strptime(dateformat, "%d-%m-%Y").date()
        except ValueError:
            # Nếu không thể chuyển đổi, in ra thông báo lỗi và bỏ qua sheet này
            print(f"Tên sheet '{sheet_name}' không đúng định dạng ngày. Sheet này sẽ bị bỏ qua.")
            continue

        df = excel_file.parse(sheet_name)
        patient_codes = df.iloc[2:, 1]  # Đọc cột mã bệnh nhân từ dòng thứ 3
        invoices_id = df.iloc[2:, 2]  # Đọc cột mã hóa đơn từ dòng thứ 3

        # Chỉ lấy những dòng là mã bệnh nhân
        valid_patient_codes = []
        valid_patient_invoices =[]
        for a, b in zip(patient_codes, invoices_id):
            code_str = str(a)  # Convert to string
            invoicestr = str(b)  # Convert to string
            if code_str[0].isdigit():
                if invoicestr != "nan":
                    valid_patient_codes.append(code_str)
                    valid_patient_invoices.append(invoicestr)                   
        
        # Nếu không có danh sách mã bệnh nhân hợp lệ thì next
        if not valid_patient_codes:
            continue

        # Tạo model
        model = {
            'date': date,
            'patient_codes': valid_patient_codes,
            'invoices_id': valid_patient_invoices
        }
        listmodel.append(model)

    start_script_thread_Invoice(listmodel, directory)

def run_script_Invoice(listmodels,directory,terminal_text):
    global pathExcel

    # Hàm để hiển thị thông tin lên terminal
    def log_terminal(message):
        terminal_text.insert(tk.END,message + "\n")
        terminal_text.see(tk.END)  # Scroll to the end
        terminal_text.update_idletasks()

    #duyệt danh sách model
    for model in listmodels:
        dateKham = model['date']
        dateKham = dateKham.strftime("%Y-%m-%d")
        patient_codes_list = model['patient_codes']
        ticket_list = model['tickets_id']
        Invoice_result_list = []
        #csv file
        csv_filename = os.path.join(directory, f"Invoice_Data_{dateKham}.csv")
        # CSV file header
        csv_header = ["STT","Mã bệnh nhân","Mã hóa đơn" ,"Bảo hiểm y tế", "Thời gian tạo", "Tên thu ngân", "Loại dịch vụ khám", "Loại hóa đơn", "Loại bệnh nhân", "Hình thức thanh toán", "pay_receipt_id", "Mã đợt khám", "Số hóa đơn", "Số lưu trữ", "Tổng tiền", "Bệnh nhân trả", "Tiền hoàn lại","Hoàn trả hóa đơn", "Xóa", "Ghi chú", "IDMapping"]
         # Kiểm tra nếu file CSV đã tồn tại và đếm số dòng
        start_stt = 0
        start_code = ""
        idmapcurrent = 0
        if os.path.exists(csv_filename):
            with open(csv_filename, mode='r', encoding='utf-8-sig') as file:
                reader = csv.reader(file)
                next(reader) #bỏ qua dòng 1
                for row in reader:
                    if len(row) > 1:
                        start_stt+=1
                        start_code = row[22]
        if start_code != "":
            idmapcurrent = int(start_code)
        demstt = start_stt
        breakSTT = 1
        for pt,tl in zip(patient_codes_list[idmapcurrent:],ticket_list[idmapcurrent:]):
            success = False
            errorDrug = 0
            errorLogWeb = 0 
            while not success:
                try:
                    if breakSTT == 1 and errorDrug == 0:
                        #chrome settings
                        chrome_options = webdriver.ChromeOptions()
                        chrome_options.add_experimental_option("prefs", {
                            "download.default_directory": directory,
                            "profile.default_content_setting_values.automatic_downloads": 1,
                            "download.prompt_for_download": False,
                            "profile.default_content_settings.popups": 0,
                            "safebrowsing.enabled": "false",
                            "safebrowsing.disable_download_protection": True
                        })
                        ascii_nd2 = """
........................Đang kết nối lại website..............................
                        """
                        log_terminal(ascii_nd2)
                        prefs = {"credentials_enable_service": False,
                            "profile.password_manager_enabled": False}
                        chrome_options.add_experimental_option("prefs", prefs)
                        chrome_options.add_argument("--headless")  # Chạy trình duyệt trong chế độ headless
                        chrome_options.add_argument("--disable-gpu")  # Tăng tốc độ trên các hệ điều hành không có GPU
                        chrome_options.add_argument("--window-size=1920x1080")  # Thiết lập kích thước cửa sổ mặc định
                        chrome_options.add_argument("--no-sandbox")
                        chrome_options.add_argument("--disable-extensions")
                        chrome_options.add_argument("--disable-infobars")
                        chrome_options.add_argument("--blink-settings=imagesEnabled=false")
                        driver = webdriver.Chrome(options=chrome_options)
                        url = "http://192.168.0.77/dist/#!/login"
                        try:
                            driver.get(url)
                        except Exception as e:
                            errorLogWeb += 1
                            log_terminal(f"Lỗi khi truy cập URL {url}: {e}")
                            if errorLogWeb >= 10:
                                log_terminal("Lỗi server, dừng chương trình.")
                                return
                            time.sleep(3)  # Chờ một chút trước khi thử lại
                            continue

                        errorLogWeb = 0  # Reset biến đếm lỗi khi truy cập URL thành công
                        time.sleep(3)
                        log_terminal(".........................Duyệt WEBSITE thành công.................................")
                        # login
                        form = driver.find_element(By.NAME, "LoginForm")
                        username = driver.find_element(By.NAME, 'username')
                        password = driver.find_element(By.NAME, 'password')
                        username.send_keys('quyen.ngoq')
                        password.send_keys('74777477')
                        # ấn nút login
                        form.submit()
                        time.sleep(2)
                        # Chọn
                        liElement = driver.find_element(By.XPATH,
                                                        "//li[contains(@class, 'nav-item pl-2 pr-3 d-lg-none d-xl-block') and contains(@permission, "
                                                        "'patient_info_payment')]")
                        liElement.click()
                        time.sleep(1)
                    cod = str(pt)
                    patientcode = driver.find_element(By.NAME, 'patient_code_qr')
                    patientcode.clear()
                    patientcode.send_keys(cod)
                    time.sleep(0.1)
                    btnFind = driver.find_element(By.XPATH, "//button[contains(text(), 'Tìm')]")
                    btnFind.click()
                    time.sleep(0.1)
                    btnThuoc = driver.find_element(By.XPATH, "//a[contains(text(), 'Khám Ngoại Trú ')]")
                    btnThuoc.click()
                    time.sleep(0.1)
                    #tìm mã toa phù hợp
                    for request in reversed(driver.requests):
                        foundDrug = False
                        if request.response:
                            if 'getInvoices_outpatient' in request.url:
                                # format cho dữ liệu về dạng json
                                try:
                                    drug_data = json.loads(request.response.body)
                                    # đọc data
                                    invoice_info = drug_data.get('data', {}).get('invoices', {})
                                    for di in invoice_info:
                                        if di.get('ticket_id') == int(tl):
                                            if foundDrug!= True:
                                                foundDrug = True 
                                                idmapcurrent += 1
                                            invoice_result = InvoiceModel()
                                            invoice_result.ticket_id = int(tl)
                                            invoice_result.patient_code = pt
                                            invoice_result.invoice_code = di.get('invoice_code')
                                            invoice_result.bhyt = di.get('bhyt')
                                            invoice_result.created_time = di.get('created_time')
                                            invoice_result.cashier_name = di.get('cashier_name')
                                            invoice_result.enum_examination_type = di.get('enum_examination_type')
                                            invoice_result.enum_invoice_type = di.get('enum_invoice_type')
                                            invoice_result.enum_patient_type = di.get('enum_patient_type')
                                            invoice_result.enum_payment_type = di.get('enum_payment_type')
                                            invoice_result.pay_receipt_id = di.get('pay_receipt_id')
                                            invoice_result.so_hd = di.get('so_hd')
                                            invoice_result.so_luu_tru = di.get('so_luu_tru')
                                            invoice_result.total_price = di.get('total_price')
                                            invoice_result.patient_price = di.get('patient_price')
                                            invoice_result.total_returned_amount = di.get('total_returned_amount')
                                            invoice_result.returned = di.get('returned')
                                            invoice_result.deleted = di.get('deleted')
                                            invoice_result.note = di.get('note')
                                            demstt+=1
                                            invoice_result.stt = demstt
                                            invoice_result.IDMapping = idmapcurrent
                                            Invoice_result_list.append(invoice_result)
                                            log_terminal(invoice_result.ExportModel())
                                        else:
                                            if foundDrug == True:
                                                #đã tìm xong r
                                                break 
                                except Exception as e:
                                    print(e)
                                    continue    
                            else:
                                continue
                        if foundDrug == True:
                            break
                    if foundDrug == False:
                        success = False
                        errorDrug += 1
                        if errorDrug >= 10:
                            log_terminal(f"Quá nhiều lỗi dữ liệu trống, bỏ qua kết thúc quá trình cào tại mã bệnh nhân {cod}")
                            app.destroy()
                        else:
                            log_terminal(f"Tìm chưa thấy hóa đơn! Đang cố thử lại...")
                    else:
                        success = True
                        if breakSTT % 100 == 0:
                            # Ghi dữ liệu tạm thời vào file CSV
                            with open(csv_filename, mode='a', newline='', encoding='utf-8') as file:
                                writer = csv.DictWriter(file, fieldnames=csv_header)
                                if file.tell() == 0:  # Nếu file rỗng, ghi header
                                    writer.writeheader()
                                writer.writerows([pt.to_dict() for pt in Invoice_result_list])
                            breakSTT = 1
                            driver.quit()
                            Invoice_result_list.clear()
                            log_terminal(".........................Ghi tạm vào file CSV........................................")
                        else:
                            breakSTT = breakSTT + 1
                except:
                    log_terminal(f"Có lỗi xảy ra! Đang cố thử lại...")
                    success = False
        # Ghi dữ liệu còn lại vào file CSV nếu có
        if Invoice_result_list:
            with open(csv_filename, mode='a', newline='', encoding='utf-8') as file:
                writer = csv.DictWriter(file, fieldnames=csv_header)
                if file.tell() == 0:
                    writer.writeheader()
                writer.writerows([pt.to_dict() for pt in Invoice_result_list])
        # Thoát trình duyệt 
        driver.quit()
        # Chuyển dữ liệu từ file CSV sang file Excel
        try:
            df = pd.read_csv(csv_filename, encoding='utf-8-sig',dtype={'Mã bệnh nhân': str,'Số hóa đơn': str})
            output_filename = os.path.join(directory, f"Invoice_Data_{dateKham}.xlsx")
            with pd.ExcelWriter(output_filename, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name=f'{dateKham}')

            # Định dạng file Excel
            wb = load_workbook(output_filename)
            ws = wb.active

            # Định dạng header
            header_font = Font(bold=True)
            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Tự động điều chỉnh độ rộng cột
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            
            # Định dạng cột Mã bệnh nhân là Text
            for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
                for cell in row:
                    cell.number_format = '@'
            
            # Định dạng cột số hóa đơn là Text
            for row in ws.iter_rows(min_row=2, min_col=13, max_col=13):
                for cell in row:
                    cell.number_format = '@'
            wb.save(output_filename)
            os.remove(csv_filename)  # Xóa file CSV sau khi chuyển sang Excel
            log_terminal(f"Quá trình thu thập dữ liệu hóa đơn ngày {dateKham} thành công!")
        except Exception as e:
            messagebox.showerror(title="Lỗi", message=f"Lỗi quá trình tạo file excel không thành công! {e}")  
    
    # Đóng cửa sổ terminal và hiển thị lại cửa sổ chính
    app.deiconify()

def run_script_InvoiceDetail(listmodels,directory,terminal_text):
    global pathExcel

    # Hàm để hiển thị thông tin lên terminal
    def log_terminal(message):
        terminal_text.insert(tk.END,message + "\n")
        terminal_text.see(tk.END)  # Scroll to the end
        terminal_text.update_idletasks()

    #duyệt danh sách model
    for model in listmodels:
        dateKham = model['date']
        dateKham = dateKham.strftime("%Y-%m-%d")
        patient_codes_list = model['patient_codes']
        invoices_list = model['invoices_id']
        InvoiceDetail_result_list = []
        #csv file
        csv_filename = os.path.join(directory, f"InvoiceDetail_Data_{dateKham}.csv")
        # CSV file header
        csv_header = ["STT",
                      "Mã hóa đơn" ,
                      "pay_payment_item_id", 
                      "pay_receipt_id", 
                      "ID bệnh nhân", 
                      "ID thu ngân", 
                      "Tên thu ngân", 
                      "Ngày thu", 
                      "Người thu", 
                      "Tên bác sĩ", 
                      "Thời gian tạo", 
                      "Mã hóa đơn điện tử", 
                      "Trạng thái hóa đơn điện tử", 
                      "Loại hóa đơn", 
                      "Phân loại", 
                      "Loại thanh toán",
                      "Ngày hóa đơn", 
                      "Thông tin hóa đơn", 
                      "Đã thu", 
                      "Đã xóa", 
                      "Hoàn trả", 
                      "refund_date", 
                      "refund_amount", 
                      "refund_invoice_code", 
                      "refund_pay_receipt_id", 
                      "new_invoice_code", 
                      "ID dịch vụ", 
                      "Mã dịch vụ", 
                      "Tên dịch vụ", 
                      "Tên bảo hiểm", 
                      "Loại dịch vụ", 
                      "Đơn vị tính", 
                      "Đơn giá", 
                      "Số lượng", 
                      "Total", 
                      "total_dichvu", 
                      "total_nutrition", 
                      "total_price", 
                      "patient_price", 
                      "total_returned_amount", 
                      "tam_ung", 
                      "phai_thu", 
                      "phai_tra", 
                      "so_hd", 
                      "so_luu_tru", 
                      "shift", 
                      "Khu vực", 
                      "Ghi chú", 
                      "IDMapping"]
         # Kiểm tra nếu file CSV đã tồn tại và đếm số dòng
        start_stt = 0
        start_code = ""
        idmapcurrent = 0
        if os.path.exists(csv_filename):
            with open(csv_filename, mode='r', encoding='utf-8-sig') as file:
                reader = csv.reader(file)
                next(reader) #bỏ qua dòng 1
                for row in reader:
                    if len(row) > 1:
                        start_stt+=1
                        start_code = row[49]
        if start_code != "":
            idmapcurrent = int(start_code)
        demstt = start_stt
        breakSTT = 1
        for pt,il in zip(patient_codes_list[idmapcurrent:],invoices_list[idmapcurrent:]):
            success = False
            errorDrug = 0
            errorLogWeb = 0 
            while not success:
                try:
                    if breakSTT == 1 and errorDrug == 0:
                        #chrome settings
                        chrome_options = webdriver.ChromeOptions()
                        chrome_options.add_experimental_option("prefs", {
                            "download.default_directory": directory,
                            "profile.default_content_setting_values.automatic_downloads": 1,
                            "download.prompt_for_download": False,
                            "profile.default_content_settings.popups": 0,
                            "safebrowsing.enabled": "false",
                            "safebrowsing.disable_download_protection": True
                        })
                        ascii_nd2 = """
........................Đang kết nối lại website..............................
                        """
                        log_terminal(ascii_nd2)
                        prefs = {"credentials_enable_service": False,
                            "profile.password_manager_enabled": False}
                        chrome_options.add_experimental_option("prefs", prefs)
                        chrome_options.add_argument("--headless")  # Chạy trình duyệt trong chế độ headless
                        chrome_options.add_argument("--disable-gpu")  # Tăng tốc độ trên các hệ điều hành không có GPU
                        chrome_options.add_argument("--window-size=1920x1080")  # Thiết lập kích thước cửa sổ mặc định
                        chrome_options.add_argument("--no-sandbox")
                        chrome_options.add_argument("--disable-extensions")
                        chrome_options.add_argument("--disable-infobars")
                        chrome_options.add_argument("--blink-settings=imagesEnabled=false")
                        driver = webdriver.Chrome(options=chrome_options)
                        url = "http://192.168.0.77/dist/#!/login"
                        try:
                            driver.get(url)
                        except Exception as e:
                            errorLogWeb += 1
                            log_terminal(f"Lỗi khi truy cập URL {url}: {e}")
                            if errorLogWeb >= 10:
                                log_terminal("Lỗi server, dừng chương trình.")
                                return
                            time.sleep(3)  # Chờ một chút trước khi thử lại
                            continue

                        errorLogWeb = 0  # Reset biến đếm lỗi khi truy cập URL thành công
                        time.sleep(3)
                        log_terminal(".........................Duyệt WEBSITE thành công.................................")
                        # login
                        form = driver.find_element(By.NAME, "LoginForm")
                        username = driver.find_element(By.NAME, 'username')
                        password = driver.find_element(By.NAME, 'password')
                        username.send_keys('quyen.ngoq')
                        password.send_keys('74777477')
                        # ấn nút login
                        form.submit()
                        time.sleep(2)
                        # Chọn
                        liElement = driver.find_element(By.XPATH,
                                                        "//li[contains(@class, 'nav-item pl-2 pr-3 d-lg-none d-xl-block') and contains(@permission, "
                                                        "'patient_info_payment')]")
                        liElement.click()
                        time.sleep(1)
                    try:
                        passForm = driver.find_element(By.XPATH, "//button[contains(text(), 'Đóng')]")
                        passForm.click()
                        time.sleep(0.1)
                    except:
                        time.sleep(0.1)
                    cod = str(pt)
                    patientcode = driver.find_element(By.NAME, 'patient_code_qr')
                    patientcode.clear()
                    patientcode.send_keys(cod)
                    time.sleep(0.1)
                    btnFind = driver.find_element(By.XPATH, "//button[contains(text(), 'Tìm')]")
                    btnFind.click()
                    time.sleep(0.1)
                    btnThuoc = driver.find_element(By.XPATH, "//a[contains(text(), 'Khám Ngoại Trú ')]")
                    btnThuoc.click()
                    time.sleep(0.1)
                    btnInvoice = driver.find_element(By.XPATH, f"//a[contains(text(), '{str(il)}')]")
                    btnInvoice.click()
                    time.sleep(0.5)
                    #tìm mã toa phù hợp
                    for request in reversed(driver.requests):
                        foundInvoice = False
                        emptyInvoice = False
                        if request.response:
                            if 'getOutPatientInvoiceDetail' in request.url:
                                # format cho dữ liệu về dạng json
                                try:
                                    drug_data = json.loads(request.response.body)
                                    # đọc data
                                    invoice_info = drug_data.get('data', {})
                                    if len(invoice_info) > 0:
                                        foundInvoice = True
                                        for di in invoice_info: 
                                            idmapcurrent += 1
                                            invoice_result = InvoiceDetailModel()
                                            invoice_result.invoice_code = il
                                            invoice_result.pay_payment_item_id = di.get('pay_payment_item_id')
                                            invoice_result.pay_receipt_id = di.get('pay_receipt_id')
                                            invoice_result.patient_id = di.get('patient_id')
                                            invoice_result.cashier_id = di.get('cashier_id')
                                            invoice_result.cashier_name = di.get('cashier_name')
                                            invoice_result.ngay_thu = di.get('ngay_thu')
                                            invoice_result.nguoi_thu = di.get('nguoi_thu')
                                            if di.get('doctor_name') == None:
                                                invoice_result.doctor_name = ""
                                            else:
                                                invoice_result.doctor_name = di.get('doctor_name')
                                            invoice_result.created_time = di.get('created_time')
                                            invoice_result.e_invoice_id = di.get('e_invoice_id')
                                            invoice_result.enum_invoice_type = di.get('enum_invoice_type')
                                            invoice_result.enum_item_type = di.get('enum_item_type')
                                            invoice_result.enum_payment_type = di.get('enum_payment_type')
                                            invoice_result.hd_date = di.get('hd_date')
                                            invoice_result.hd_info = di.get('hd_info')
                                            invoice_result.da_thu = di.get('da_thu')
                                            invoice_result.deleted = di.get('deleted')
                                            invoice_result.returned = di.get('returned')
                                            invoice_result.refund_date = di.get('refund_date')
                                            invoice_result.refund_amount = di.get('refund_amount')
                                            invoice_result.refund_invoice_code = di.get('refund_invoice_code')
                                            invoice_result.refund_pay_receipt_id = di.get('refund_pay_receipt_id')
                                            invoice_result.new_invoice_code = di.get('new_invoice_code')
                                            invoice_result.service_id = di.get('service_id')
                                            invoice_result.service_name = di.get('service_name')
                                            invoice_result.insurance_name = di.get('insurance_name')
                                            invoice_result.item_type = di.get('item_type')
                                            invoice_result.unit = di.get('unit')
                                            invoice_result.unit_price = di.get('unit_price')
                                            invoice_result.quantity = di.get('quantity')
                                            invoice_result.total = di.get('total')
                                            invoice_result.total_dichvu = di.get('total_dichvu')
                                            invoice_result.total_nutrition = di.get('total_nutrition')
                                            invoice_result.total_price = di.get('total_price')
                                            invoice_result.patient_price = di.get('patient_price')
                                            invoice_result.total_returned_amount = di.get('total_returned_amount')
                                            invoice_result.tam_ung = di.get('tam_ung')
                                            invoice_result.phai_thu = di.get('phai_thu')
                                            invoice_result.phai_tra = di.get('phai_tra')
                                            invoice_result.so_hd = di.get('so_hd')
                                            invoice_result.so_luu_tru = di.get('so_luu_tru')
                                            invoice_result.shift = di.get('shift')
                                            invoice_result.zone = di.get('zone')
                                            invoice_result.note = di.get('note')
                                            demstt+=1
                                            invoice_result.stt = demstt
                                            invoice_result.IDMapping = idmapcurrent
                                            InvoiceDetail_result_list.append(invoice_result)
                                            log_terminal(invoice_result.ExportModel())
                                        if foundInvoice == True:
                                            break
                                        else:
                                            emptyInvoice = True
                                            break
                                except Exception as e:
                                    print(e)
                                    continue    
                            else:
                                continue
                        if foundInvoice == True or emptyInvoice == True:
                            try:
                                passForm = driver.find_element(By.XPATH, "//button[contains(text(), 'Đóng')]")
                                passForm.click()
                                time.sleep(0.1)
                            except:
                                time.sleep(0.1)
                            break
                    if emptyInvoice == False:
                        if foundInvoice == False:
                            success = False
                            errorDrug += 1
                            if errorDrug >= 10:
                                log_terminal(f"Quá nhiều lỗi dữ liệu trống, bỏ qua kết thúc quá trình cào tại mã bệnh nhân {cod}")
                                app.destroy()
                            else:
                                log_terminal(f"Tìm chưa thấy chi tiết hóa đơn! Đang cố thử lại...")
                        else:
                            success = True
                            if breakSTT % 100 == 0:
                                # Ghi dữ liệu tạm thời vào file CSV
                                with open(csv_filename, mode='a', newline='', encoding='utf-8') as file:
                                    writer = csv.DictWriter(file, fieldnames=csv_header)
                                    if file.tell() == 0:  # Nếu file rỗng, ghi header
                                        writer.writeheader()
                                    writer.writerows([pt.to_dict() for pt in InvoiceDetail_result_list])
                                breakSTT = 1
                                driver.quit()
                                InvoiceDetail_result_list.clear()
                                log_terminal(".........................Ghi tạm vào file CSV........................................")
                            else:
                                breakSTT = breakSTT + 1
                    else:
                        success = False
                        errorDrug += 1
                        log_terminal(".........................Hóa đơn rỗng!!!........................................")
                except:
                    log_terminal(f"Có lỗi xảy ra! Đang cố thử lại...")
                    success = False
        # Ghi dữ liệu còn lại vào file CSV nếu có
        if InvoiceDetail_result_list:
            with open(csv_filename, mode='a', newline='', encoding='utf-8') as file:
                writer = csv.DictWriter(file, fieldnames=csv_header)
                if file.tell() == 0:
                    writer.writeheader()
                writer.writerows([pt.to_dict() for pt in InvoiceDetail_result_list])
        # Thoát trình duyệt 
        driver.quit()
        # Chuyển dữ liệu từ file CSV sang file Excel
        try:
            df = pd.read_csv(csv_filename, encoding='utf-8-sig', dtype={'so_hd': str})
            output_filename = os.path.join(directory, f"InvoiceDetail_Data_{dateKham}.xlsx")
            with pd.ExcelWriter(output_filename, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name=f'{dateKham}')

            # Định dạng file Excel
            wb = load_workbook(output_filename)
            ws = wb.active

            # Định dạng header
            header_font = Font(bold=True)
            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Tự động điều chỉnh độ rộng cột
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            
            # Định dạng cột Mã bệnh nhân là Text
            for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
                for cell in row:
                    cell.number_format = '@'
            
            wb.save(output_filename)
            os.remove(csv_filename)  # Xóa file CSV sau khi chuyển sang Excel
            log_terminal(f"Quá trình thu thập dữ liệu chi tiết hóa đơn ngày {dateKham} thành công!")
        except Exception as e:
            messagebox.showerror(title="Lỗi", message=f"Lỗi quá trình tạo file excel không thành công! {e}")  
    
    # Đóng cửa sổ terminal và hiển thị lại cửa sổ chính
    app.deiconify()

def validate_input():
    global radio_var
    if radio_var.get() == 0:
        if pathExcel == "":
            return False
        return True
    else:
        if pathExcel == "":
            messagebox.showerror(title= "Đã có lỗi", message="Vui lòng chọn file excel trước!")
            return False
        else:
            fName = os.path.basename(pathExcel)
            if fName.startswith("Invoice_Data"):
                return True
            messagebox.showerror(title= "File không hợp lệ!", message="Vui lòng chọn file excel dữ liệu hóa đơn!")
            return False

def cancel_Excel_file():
    global pathExcel, drop
    if pathExcel == '':
        messagebox.showerror(title="Thất bại!", message="Chưa chọn file nào!")
    else:
        pathExcel = ''
        drop.configure(text="Ấn để chọn file Excel!")
        # Cập nhật fg_color của nút
        drop.configure(fg_color="#FF9292")
        # Bật enable nút
        drop.configure(state="enable")

def select_Excel_file():
    global pathExcel, drop
    # Hiển thị hộp thoại chọn file PDF
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if file_path:
        # Lấy đường dẫn file
        pathExcel = file_path
        # Lấy tên file từ đường dẫn
        file_name = os.path.basename(file_path)
        # Cập nhật text của nút với tên file
        drop.configure(text=file_name)
        # Cập nhật fg_color của nút
        drop.configure(fg_color="#5DBD60")
        # Tắt enable nút
        drop.configure(state="disabled")

def center_window(window, width=600, height=440):
    # Lấy kích thước màn hình
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    # Tính toán tọa độ x và y để cửa sổ ở giữa màn hình
    x = (screen_width / 2) - (width / 2)
    y = (screen_height / 2) - (height / 2)
    window.geometry(f'{width}x{height}+{int(x)}+{int(y)}')
#app
def run_secondary_interface(main_app):
    global run_button, app, drop, radio_var
    
    radio_var = customtkinter.IntVar(value=0)

    app = customtkinter.CTkToplevel(main_app)
    app.title("Lấy dữ liệu hóa đơn")
    center_window(app)
    icon_path = os.path.abspath("resource\crawlLogo.ico")
    app.iconbitmap(icon_path)

    def on_closing():
        main_app.deiconify()  # Hiển thị lại cửa sổ chính khi đóng cửa sổ mới
        app.destroy()

    app.protocol("WM_DELETE_WINDOW", on_closing)

    imgBG = ImageTk.PhotoImage(Image.open(r"resource\backgroundInvoice.jpg"))
    l1 = customtkinter.CTkLabel(master=app, image=imgBG)
    l1.pack()

    frame = customtkinter.CTkFrame(master=l1, width=320, height=300, corner_radius=15)
    frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

    l2 = customtkinter.CTkLabel(master=frame, text="Select configuration", font=('Century Gothic', 20))
    l2.place(x=70, y=45)

    btnInvoice = customtkinter.CTkRadioButton(master=frame, variable=radio_var, value=0, text="Dữ liệu hóa đơn", text_color="#B3271C")
    btnInvoice.place(x= 10, y= 90)

    btnInvoiceDetail = customtkinter.CTkRadioButton(master=frame, variable=radio_var, value=1, text="Dữ liệu dịch vụ", text_color="#B3271C")
    btnInvoiceDetail.place(x= 170, y= 90)

    drop = customtkinter.CTkButton(master=frame, command=select_Excel_file, width=300, height=50, corner_radius=15, fg_color="#FF9292", hover_color="#FFB6B6", text_color="#000000", text="Ấn để chọn file Excel!", font=('Tahoma', 13))
    drop.place(x=10, y=130)

    cancel_button = customtkinter.CTkButton(master=frame, text="Hủy bỏ", command=cancel_Excel_file, font=('Tahoma', 13), fg_color="#B3271C", hover_color="#FF3D4D")
    cancel_button.place(x=10, y=200)

    run_button = customtkinter.CTkButton(master=frame, text="Thực thi", command=handle_file_excel, font=('Tahoma', 13), fg_color="#005369", hover_color="#008097")
    run_button.place(x=170, y=200)

    app.mainloop()